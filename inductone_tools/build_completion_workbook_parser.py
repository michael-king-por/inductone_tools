"""
Parser for the OPS-BLD-F01 Builder Serial Workbook.

The workbook has a known structure (Builder Input sheet, column A labels,
column B values, sections delimited by 'SECTION X - ...' headers). This
module reads an uploaded workbook and returns structured data the
completion-upload flow can persist.

Parser philosophy:
  - Find labels by string match in column A, not by hardcoded row numbers.
    If the template ever has rows inserted, the parser still works.
  - Section header rows (containing 'SECTION ') are not treated as data.
  - Empty cells produce rows with empty serial_number — per Q4 in the
    design discussion: 'make a row for everything, including missing,
    and populate the component with an empty serial number so it's
    clear what's missing'.
  - Section A (Build Metadata) and Section C-Attestation are returned
    separately from the component-serial rows, so callers can decide
    how to use them.
"""

import io

from openpyxl import load_workbook


# Labels in column A that are NOT component serials, even though they
# look like data rows. These go to the metadata bucket instead.
METADATA_LABELS = {
    "InductOne Serial Number (IND-####)",
    "Build Date",
    "Builder Organization",
    "Builder Point of Contact",
    "Builder Point of Contact Email",
}

ATTESTATION_LABELS = {
    "Builder Signature (Typed Full Name)",
    "Date",
    "I confirm all entries are accurate (YES/NO)",
}


class WorkbookParseError(Exception):
    """Raised when the workbook cannot be parsed at all (wrong format,
    missing sheet, corrupted file). Caller should reject the upload
    and tell the user the workbook structure is wrong."""
    pass


def parse_builder_workbook(file_bytes):
    """
    Parse an uploaded OPS-BLD-F01 workbook.

    Args:
        file_bytes: raw bytes of the .xlsx file.

    Returns:
        dict with keys:
            metadata: dict[str, str] — Section A values keyed by label
            attestation: dict[str, str] — attestation values keyed by label
            components: list[dict] — one per non-empty row in Sections B-I,
                                     each with keys: component_label, serial_number
            warnings: list[str] — non-fatal anomalies (unknown labels, etc.)

    Raises:
        WorkbookParseError on structural failure.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise WorkbookParseError(
            f"Could not open workbook: {e}. "
            "Confirm the file is a valid .xlsx and follows the OPS-BLD-F01 template."
        )

    if "Builder Input" not in wb.sheetnames:
        raise WorkbookParseError(
            "Workbook is missing the required 'Builder Input' sheet. "
            "Confirm the builder used the unmodified OPS-BLD-F01 template."
        )

    ws = wb["Builder Input"]

    metadata = {}
    attestation = {}
    components = []
    warnings = []

    # Walk the sheet row-by-row. Skip section header rows entirely
    # (they have 'SECTION ' in column A). Skip totally-empty rows.
    for row_idx in range(1, ws.max_row + 1):
        label_cell = ws.cell(row=row_idx, column=1).value
        value_cell = ws.cell(row=row_idx, column=2).value

        if label_cell is None:
            continue

        label = str(label_cell).strip()

        if not label:
            continue

        # Section header row — skip.
        if label.upper().startswith("SECTION "):
            continue

        value = _normalize_value(value_cell)

        if label in METADATA_LABELS:
            metadata[label] = value
            continue

        if label in ATTESTATION_LABELS:
            attestation[label] = value
            continue

        # Anything else in column A is a component serial label.
        # Even if empty, record the row — operator should see what
        # was left blank.
        components.append({
            "component_label": label,
            "serial_number": value,
        })

    if not components:
        # If we walked the whole sheet and found zero component rows,
        # something is structurally wrong (e.g., the builder pasted
        # data into the wrong sheet or replaced the template).
        raise WorkbookParseError(
            "Workbook contained no component serial rows. "
            "Confirm the builder used the unmodified OPS-BLD-F01 "
            "template and filled in the Builder Input sheet."
        )

    return {
        "metadata": metadata,
        "attestation": attestation,
        "components": components,
        "warnings": warnings,
    }


def _normalize_value(value):
    """Convert a cell value to a stripped string. Empty cells, None,
    and whitespace-only strings all become ''."""
    if value is None:
        return ""
    s = str(value).strip()
    # Treat common 'no value' tokens as empty.
    if s.upper() in ("NA", "N/A", "NONE", "-"):
        return ""
    return s


def validate_workbook_against_build(parsed, build_system_serial):
    """
    Cross-check parsed workbook metadata against what we know about
    the Build. Returns a list of warning strings — these are non-fatal
    (the upload still goes through) but they get surfaced on the
    completion record for ops to review.

    Specifically checks: the IND serial in the workbook should match
    the Build's system_serial. Since we now pre-fill the workbook with
    the serial before sending to the builder (per the workbook
    generation patch), a mismatch means the builder edited a field
    they shouldn't have.
    """
    warnings = []

    if not build_system_serial:
        return warnings  # Nothing to compare against.

    wb_serial = parsed.get("metadata", {}).get(
        "InductOne Serial Number (IND-####)", ""
    ).strip()

    if not wb_serial:
        warnings.append(
            f"Workbook Section A is missing the InductOne Serial Number. "
            f"Build has '{build_system_serial}'. The serial was pre-filled "
            f"in the workbook when released; verify the builder did not "
            f"clear it."
        )
        return warnings

    if wb_serial != build_system_serial:
        warnings.append(
            f"Workbook serial '{wb_serial}' does not match Build's "
            f"allocated serial '{build_system_serial}'. The builder may "
            f"have stenciled the wrong number on the unit, or edited the "
            f"pre-filled field. Investigate before accepting."
        )

    return warnings
