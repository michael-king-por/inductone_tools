import csv
import io
import os
from tempfile import NamedTemporaryFile

import frappe
from frappe.utils import now
from frappe.utils.file_manager import save_file
from frappe import _

from openpyxl import load_workbook

from .bom_export import build_configured_rows

@frappe.whitelist()
def check_builder_release_readiness(build_name: str):
    """
    Read-only readiness check for builder release.
    
    Validates that everything required to release is in place.
    Does NOT generate any artifacts. Does NOT modify any state.
    Returns a structured result with ready/missing/warnings.
    """
    if not build_name:
        frappe.throw(_("build_name is required."))
    
    build = frappe.get_doc("InductOne Build", build_name)
    
    missing = []
    warnings = []
    
    # ---- Build-level requirements ----
    
    if not getattr(build, "builder_supplier", None):
        missing.append("Builder (Supplier) is not set on the build.")
    
    if not getattr(build, "top_item", None):
        missing.append("Top Item is not set on the build.")
    
    top_bom = getattr(build, "top_bom", None) or getattr(build, "bom", None)
    if not top_bom:
        missing.append("Top BOM is not set on the build.")

    # System serial must be allocated before release; the builder needs
    # to know what to stencil on the unit, and the release package
    # (CO PDF, builder instructions) won't carry a serial if none exists.
    if not getattr(build, "system_serial", None):
        missing.append(
            "System serial has not been allocated. "
            "Click 'Allocate Serial' on the build to assign one from the "
            "builder's tranche before releasing."
        )

    snapshot_name = _resolve_snapshot_name(build)
    if not snapshot_name:
        missing.append("No snapshot has been generated for this build.")
    
    co_name = _resolve_configuration_order_name(build)
    if not co_name:
        missing.append("No Configuration Order has been generated for this build.")
    
    # ---- CO-level requirements ----
    
    co = None
    if co_name:
        try:
            co = frappe.get_doc("InductOne Configuration Order", co_name)
        except Exception:
            missing.append(f"Configuration Order {co_name} could not be loaded.")
    
    if co and snapshot_name:
        co_snapshot = getattr(co, "snapshot", None)
        if co_snapshot and co_snapshot != snapshot_name:
            missing.append(
                f"Configuration Order points to snapshot {co_snapshot}, "
                f"but build is pointing to snapshot {snapshot_name}. "
                f"Regenerate the Configuration Order from the current snapshot."
            )

    # CO and Build must agree on the system serial. The release gate
    # in release_to_builder_now self-heals this case (stamps the CO from
    # the Build) but reporting it here gives the user a chance to fix it
    # via the normal flow — clicking Allocate Serial again will
    # re-propagate to the CO.
    if co:
        build_serial = getattr(build, "system_serial", None)
        co_serial = getattr(co, "system_serial", None)
        if build_serial and co_serial and build_serial != co_serial:
            missing.append(
                f"System serial mismatch: Build has '{build_serial}' "
                f"but Configuration Order has '{co_serial}'. "
                f"This indicates a data drift — investigate manually."
            )
        elif build_serial and not co_serial:
            warnings.append(
                f"Build has system serial '{build_serial}' but the "
                f"Configuration Order does not. Click 'Allocate Serial' on "
                f"the build to re-propagate, or proceed with release "
                f"(the release will self-heal this case)."
            )
    
    if co:
        co_status = getattr(co, "co_status", None)
        if co_status and co_status not in ("Draft", "Released"):
            warnings.append(
                f"Configuration Order is in status '{co_status}'. "
                f"Releasing now will not change that state."
            )
    
    # ---- BOM Export Package requirements ----
    
    package_name = _resolve_bom_export_package_name(build, co) if co else None
    if not package_name:
        missing.append("No BOM Export Package is linked to this build.")
    else:
        try:
            package_doc = frappe.get_doc("BOM Export Package", package_name)
            if not getattr(package_doc, "output_zip", None):
                missing.append(
                    f"BOM Export Package {package_name} has not been generated. "
                    f"Open it and run generation before releasing."
                )
            elif getattr(package_doc, "status", None) != "Complete":
                warnings.append(
                    f"BOM Export Package {package_name} status is '{package_doc.status}', not 'Complete'. "
                    f"Verify it generated successfully."
                )
        except Exception as e:
            missing.append(f"BOM Export Package {package_name} could not be loaded: {str(e)}")
    
    # ---- Flat BOM check ----
    
    if co:
        flat_bom_url = _resolve_flat_bom_file_url(build, co)
        if not flat_bom_url:
            warnings.append(
                "No Flat BOM CSV is currently linked to the Configuration Order. "
                "It will be generated at release time if missing."
            )
    
    # ---- Already released? ----
    
    if getattr(build, "build_status", None) == "RELEASED_TO_BUILDER":
        warnings.append(
            "This build has already been released to the builder. "
            "Releasing again will regenerate artifacts."
        )
    
    ready = len(missing) == 0
    
    return {
        "ok": True,
        "ready": ready,
        "build": build_name,
        "configuration_order": co_name,
        "bom_export_package": package_name,
        "missing": missing,
        "warnings": warnings,
    }

@frappe.whitelist()
def generate_builder_release_bundle(build_name: str, package_name: str = None):
    """
    Prepare the builder handoff package using the Configuration Order document list
    as the authoritative package index.

    IMPORTANT:
    This does NOT build a giant nested ZIP. Instead it creates a lightweight
    manifest artifact and ensures the Configuration Order document list is aligned
    to the current release basis.

    The Configuration Order document list is the builder package.
    """
    build = frappe.get_doc("InductOne Build", build_name)

    co_name = _resolve_configuration_order_name(build)
    snapshot_name = _resolve_snapshot_name(build)
    top_bom = getattr(build, "top_bom", None) or getattr(build, "bom", None)

    if not co_name:
        frappe.throw("InductOne Build must have a linked Configuration Order to prepare builder handoff.")

    if not snapshot_name:
        frappe.throw("InductOne Build must have a selected/latest snapshot to prepare builder handoff.")

    if not top_bom:
        frappe.throw("InductOne Build must have Top BOM to prepare builder handoff.")

    co = frappe.get_doc("InductOne Configuration Order", co_name)

    package_doc = None
    if package_name:
        package_doc = frappe.get_doc("BOM Export Package", package_name)
    else:
        resolved_package_name = _resolve_bom_export_package_name(build, co)
        if resolved_package_name:
            package_doc = frappe.get_doc("BOM Export Package", resolved_package_name)

    flat_bom_file_url = _resolve_flat_bom_file_url(build, co)

    if not package_doc:
        frappe.throw("No BOM Export Package is linked/resolved for this build.")
    if not getattr(package_doc, "output_zip", None):
        frappe.throw(
            f"BOM Export Package {package_doc.name} has not been generated yet. "
            f"Open the BOM Export Package and run generation there before preparing builder handoff."
        )

    manifest_text, manifest_json = _build_builder_release_manifest(
        build=build,
        configuration_order=co,
        package_doc=package_doc,
        flat_bom_file_url=flat_bom_file_url,
        snapshot_name=snapshot_name,
        top_bom=top_bom,
    )

    manifest_fname = f"{build.name}_builder_release_manifest_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.txt"
    saved = save_file(
        fname=manifest_fname,
        content=manifest_text.encode("utf-8"),
        dt="InductOne Build",
        dn=build.name,
        is_private=1,
    )

    _set_if_present(build, ["builder_release_bundle", "builder_package", "released_builder_package"], saved.file_url)
    _set_if_present(build, ["builder_release_manifest_json"], frappe.as_json(manifest_json, indent=2))
    _set_if_present(build, ["builder_release_generated_at", "builder_package_generated_at"], now())
    _set_if_present(build, ["builder_release_status"], "Prepared")
    build.save(ignore_permissions=True)

    # Ensure the current package components are represented on the Configuration Order
    _sync_bom_export_document_index(
        configuration_order_name=co.name,
        top_bom=top_bom,
        package_doc=package_doc,
    )

    _sync_flat_bom_document_index(
        configuration_order_name=co.name,
        build_name=build.name,
        flat_bom_file_url=flat_bom_file_url,
    )

    _sync_builder_release_document_index(
        configuration_order_name=co.name,
        build_name=build.name,
        manifest_url=saved.file_url,
        package_name=package_doc.name if package_doc else None,
        flat_bom_file_url=flat_bom_file_url,
    )

    return {
        "ok": True,
        "build": build.name,
        "configuration_order": co.name,
        "bundle_file_url": saved.file_url,
        "package_name": package_doc.name if package_doc else None,
        "flat_bom_file_url": flat_bom_file_url,
        "manifest": manifest_json,
    }


@frappe.whitelist()
def release_to_builder_now(build_name: str, package_name: str = None, note: str = None):
    """
    Release a build to the builder.
    
    Generates the manifest and workbook, syncs the CO document index,
    and stamps the build/CO status. Throws if readiness check fails.
    """
    # Defense in depth: re-validate readiness server-side before generating anything
    # Defense in depth: re-validate readiness server-side before generating anything
    readiness = check_builder_release_readiness(build_name)
    if not readiness.get("ready"):
        missing_list = "\n".join(f"  • {m}" for m in readiness.get("missing", []))
        frappe.throw(
            f"Cannot release — readiness check failed:\n\n{missing_list}"
        )

    # Serial gate: refuse to release without a system_serial on the CO.
    # The release package (CO PDF, builder instructions) must communicate
    # the stenciled serial to the builder; without it, the builder has no
    # instruction on what to apply to the unit. The assertion self-heals
    # if the parent Build has a serial but the CO doesn't (stamps it and
    # proceeds), which handles the case where someone allocated a serial
    # before this gate was added.
    from inductone_tools.serial_allocation.co_sync import assert_co_has_serial
    build = frappe.get_doc("InductOne Build", build_name)
    co_name_for_gate = _resolve_configuration_order_name(build)
    if co_name_for_gate:
        assert_co_has_serial(co_name_for_gate)

    # Generate workbook FIRST so its file_url is available when the manifest is written
    serial_result = generate_required_serial_capture_artifact(build_name)
    
    # Now generate bundle (manifest will pick up the workbook URL from the build)
    result = generate_builder_release_bundle(build_name=build_name, package_name=package_name)
    
    build = frappe.get_doc("InductOne Build", build_name)
    
    # Stamp build fields
    _set_if_present(build, ["build_status"], "RELEASED_TO_BUILDER")
    _set_if_present(build, ["released_at"], frappe.utils.now_datetime())
    _set_if_present(build, ["released_by"], frappe.session.user)
    _set_if_present(build, ["builder_release_status"], "Released")
    _set_if_present(build, ["builder_released_on", "released_to_builder_at"], frappe.utils.now_datetime())
    _set_if_present(build, ["builder_released_by", "released_to_builder_by"], frappe.session.user)
    _set_if_present(build, ["builder_release_note", "builder_release_notes"], note or "")
    _set_if_present(build, ["as_built_status"], "Pending Builder Submission")
    
    # Stamp the linked Configuration Order as Released
    co_name = _resolve_configuration_order_name(build)
    if co_name:
        try:
            co_status_field_exists = frappe.db.has_column("InductOne Configuration Order", "co_status")
        except Exception:
            co_status_field_exists = True

        if co_status_field_exists:
            frappe.db.set_value(
                "InductOne Configuration Order",
                co_name,
                {"co_status": "Released"},
            )
            frappe.db.commit()
    
    build.save(ignore_permissions=True)
    
    return {
        "ok": True,
        "build": build.name,
        "configuration_order": co_name,
        "co_status": "Released" if co_name else None,
        "bundle_file_url": result.get("bundle_file_url"),
        "serial_template_file_url": serial_result.get("template_file_url"),
        "released_at": getattr(build, "released_at", None),
        "released_by": getattr(build, "released_by", None),
    }

@frappe.whitelist()
def acknowledge_builder_release(build_name: str, acknowledgement_file_url: str = None, note: str = None):
    """
    Records the builder's acknowledgement of receipt for a released build.

    Transitions the linked Configuration Order from 'Released' to 'Awaiting Completion'
    and stamps acknowledgement metadata. Optionally attaches the acknowledgement file
    (typically a signed return of the release PDF) to the CO.
    """
    if not build_name:
        frappe.throw("build_name is required.")

    build = frappe.get_doc("InductOne Build", build_name)
    co_name = _resolve_configuration_order_name(build)

    if not co_name:
        frappe.throw(f"Build {build_name} has no linked Configuration Order to acknowledge.")

    co = frappe.get_doc("InductOne Configuration Order", co_name)

    current_status = getattr(co, "co_status", None)
    if current_status not in ("Released", "Awaiting Completion"):
        frappe.throw(
            f"Configuration Order {co_name} is in status '{current_status}'. "
            f"acknowledgement is only valid for 'Released' COs."
        )

    now_dt = frappe.utils.now_datetime()
    current_user = frappe.session.user

    update_payload = {
        "co_status": "Awaiting Completion",
        "acknowledged_at": now_dt,
        "acknowledged_by": current_user,
    }

    if acknowledgement_file_url:
        update_payload["acknowledgement_file"] = acknowledgement_file_url

    frappe.db.set_value(
        "InductOne Configuration Order",
        co_name,
        update_payload,
    )

    # Add a Document Index row referencing the acknowledgement so it lives on the CO bundle
    if acknowledgement_file_url:
        _append_or_update_document_index_row(
            parent_doctype="InductOne Configuration Order",
            parent_name=co_name,
            title=f"Builder acknowledgement - {build_name}",
            file_url=acknowledgement_file_url,
            source_type="MANUAL",
            source_name=build_name,
            sort_order=315,
            note=f"Acknowledged by {current_user} at {now_dt}" + (f" | {note}" if note else ""),
        )

    frappe.db.commit()

    return {
        "ok": True,
        "build": build_name,
        "configuration_order": co_name,
        "co_status": "Awaiting Completion",
        "acknowledged_at": now_dt,
        "acknowledged_by": current_user,
        "acknowledgement_file": acknowledgement_file_url or None,
    }

@frappe.whitelist()
def generate_required_serial_capture_artifact(build_name: str):
    """
    Generate a build-specific Excel workbook copy from the static
    OPS-BLD-F01 builder template.

    The template itself is controlled in the repo. This function creates
    a per-build workbook copy, pre-fills what is known, attaches it to the build,
    and adds/updates the Configuration Order document list row.
    """
    build = frappe.get_doc("InductOne Build", build_name)
    co_name = _resolve_configuration_order_name(build)

    if not co_name:
        frappe.throw("Latest Configuration Order is required before generating the builder serial capture workbook.")

    configuration_order = frappe.get_doc("InductOne Configuration Order", co_name)

    workbook_bytes = _build_builder_serial_workbook_bytes(build, configuration_order)
    fname = f"{build.name}_Builder_Serial_Confirmation_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.xlsx"

    saved = save_file(
        fname=fname,
        content=workbook_bytes,
        dt="InductOne Build",
        dn=build.name,
        is_private=1,
    )

    _set_if_present(build, ["required_serial_capture_file", "builder_serial_template", "serial_capture_template_file"], saved.file_url)
    _set_if_present(build, ["required_serial_capture_generated_at", "serial_capture_template_generated_at"], now())
    build.save(ignore_permissions=True)

    _append_or_update_document_index_row(
        parent_doctype="InductOne Configuration Order",
        parent_name=co_name,
        title=f"Builder Serial Capture Workbook - {build.name}",
        file_url=saved.file_url,
        source_type="MANUAL",
        source_name=build.name,
        sort_order=320,
        note=f"Build-specific workbook generated from OPS-BLD-F01 template for build {build.name}",
    )

    return {
        "ok": True,
        "build": build.name,
        "required_serial_rows": None,
        "template_file_url": saved.file_url,
    }


@frappe.whitelist()
def submit_as_built_now(build_name: str, serials_json: str = None, note: str = None):
    """
    Minimal as-built submission endpoint.
    Kept in place for later closeout work.
    """
    build = frappe.get_doc("InductOne Build", build_name)
    rows = []
    if serials_json:
        rows = frappe.parse_json(serials_json) or []

    _sync_as_built_serial_capture_table(build, rows)

    if rows:
        csv_bytes = _as_built_serial_capture_csv_bytes(rows)
        fname = f"{build.name}_as_built_serial_capture_{frappe.utils.now_datetime().strftime('%Y%m%d_%H%M%S')}.csv"
        saved = save_file(
            fname=fname,
            content=csv_bytes,
            dt="InductOne Build",
            dn=build.name,
            is_private=1,
        )
        _set_if_present(build, ["as_built_serial_capture_file", "as_built_serials_file"], saved.file_url)

    _set_if_present(build, ["as_built_status"], "Submitted")
    _set_if_present(build, ["builder_completion_note", "as_built_note", "as_built_notes"], note or "")
    _set_if_present(build, ["builder_completed_on", "as_built_submitted_on"], now())
    _set_if_present(build, ["builder_completed_by", "as_built_submitted_by"], frappe.session.user)
    build.save(ignore_permissions=True)

    return {"ok": True, "build": build.name, "submitted_rows": len(rows)}


@frappe.whitelist()
def close_build_from_as_built(build_name: str, note: str = None):
    """
    Minimal close action after manual review.
    Kept in place for later closeout work.
    """
    build = frappe.get_doc("InductOne Build", build_name)

    _set_if_present(build, ["as_built_status"], "Accepted")
    _set_if_present(build, ["build_closed_on", "closed_on", "completed_on"], now())
    _set_if_present(build, ["build_closed_by", "closed_by", "completed_by"], frappe.session.user)
    _set_if_present(build, ["closure_note", "close_note", "completion_note"], note or "")

    build.save(ignore_permissions=True)
    return {"ok": True, "build": build.name}


def _build_builder_release_manifest(build, configuration_order, package_doc, flat_bom_file_url, snapshot_name, top_bom):
    """
    System-generated audit record of the release event.
    Pure identifiers, file paths, and timestamps. No prose, no instructions.
    The Configuration Order PDF and Builder Instructions print format
    cover the human-facing content — this manifest is machine evidence only.
    """
    workbook_url = (
        getattr(build, "required_serial_capture_file", None)
        or getattr(build, "builder_serial_template", None)
        or getattr(build, "serial_capture_template_file", None)
        or ""
    )

    manifest_json = {
        "build": build.name,
        "configuration_order": configuration_order.name,
        "configured_snapshot": snapshot_name,
        "top_bom": top_bom,
        "bom_export_package": package_doc.name if package_doc else None,
        "bom_export_zip": getattr(package_doc, "output_zip", None) if package_doc else None,
        "flat_bom_file_url": flat_bom_file_url,
        "builder_workbook_url": workbook_url,
        "generated_at": now(),
        "generated_by": frappe.session.user,
    }

    lines = [
        "INDUCTONE BUILDER RELEASE MANIFEST",
        "=" * 60,
        "",
        f"Build:                {build.name}",
        f"Configuration Order:  {configuration_order.name}",
        f"Configured Snapshot:  {snapshot_name or ''}",
        f"Top BOM:              {top_bom or ''}",
        f"BOM Export Package:   {package_doc.name if package_doc else ''}",
        "",
        f"Generated by:         {frappe.session.user}",
        f"Generated at:         {now()}",
        "",
        "ARTIFACTS RELEASED",
        "-" * 60,
        f"BOM Export ZIP:       {getattr(package_doc, 'output_zip', '') if package_doc else ''}",
        f"Flat BOM CSV:         {flat_bom_file_url or ''}",
        f"Builder Workbook:     {workbook_url}",
    ]

    return "\n".join(lines), manifest_json


def _build_builder_serial_workbook_bytes(build_doc, configuration_order_doc):
    template_path = os.path.join(
        os.path.dirname(__file__),
        "builder_templates",
        "OPS-BLD-F01_Template.xlsx"
    )

    if not os.path.exists(template_path):
        frappe.throw(f"Builder serial workbook template not found at: {template_path}")

    wb = load_workbook(template_path)

    if "Builder Input" not in wb.sheetnames:
        frappe.throw("Builder serial workbook template is missing required sheet: Builder Input")

    ws = wb["Builder Input"]

    field_map = _field_label_to_cell_map(ws)

    build_date = getattr(build_doc, "released_at", None) or getattr(build_doc, "creation", None) or ""
    builder_org = getattr(build_doc, "builder_supplier", None) or getattr(configuration_order_doc, "builder_supplier", None) or ""
    builder_poc = getattr(build_doc, "builder_poc", None) or ""
    builder_poc_email = getattr(build_doc, "builder_poc_email", None) or getattr(build_doc, "builder_contact_email", None) or ""

    # The system_serial field is the canonical InductOne serial allocated
    # from the builder's tranche at release time. The two legacy fallback
    # field names are kept for forward/backward compatibility in case
    # other code paths used them at any point.
    inductone_serial = (
        getattr(build_doc, "system_serial", None)
        or getattr(build_doc, "inductone_serial_number", None)
        or getattr(build_doc, "machine_serial_number", None)
        or ""
    )

    prefills = {
        "InductOne Serial Number (IND-####)": inductone_serial,
        "Build Date": build_date,
        "Builder Organization": builder_org,
        "Builder Point of Contact": builder_poc,
        "Builder Point of Contact Email": builder_poc_email,
    }

    for label, value in prefills.items():
        cell_ref = field_map.get(label)
        if cell_ref:
            ws[cell_ref] = value

    # Optional: put some release metadata onto the Instructions sheet if present
    if "Instructions" in wb.sheetnames:
        ws2 = wb["Instructions"]
        start_row = ws2.max_row + 2
        ws2[f"A{start_row}"] = "Release Context"
        ws2[f"A{start_row + 1}"] = f"Build: {build_doc.name}"
        ws2[f"A{start_row + 2}"] = f"Configuration Order: {configuration_order_doc.name}"
        ws2[f"A{start_row + 3}"] = f"Snapshot: {_resolve_snapshot_name(build_doc) or ''}"
        ws2[f"A{start_row + 4}"] = f"Top BOM: {getattr(build_doc, 'top_bom', None) or ''}"

    with NamedTemporaryFile(suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        return tmp.read()


def _field_label_to_cell_map(ws):
    """
    Map column A labels to the editable/value cell in column B.
    Assumes the template schema uses:
      Column A = field label
      Column B = value cell
    """
    out = {}
    for row in range(1, ws.max_row + 1):
        label = ws[f"A{row}"].value
        if label and isinstance(label, str):
            out[label.strip()] = f"B{row}"
    return out


def _resolve_configuration_order_name(build_doc):
    return (
        getattr(build_doc, "latest_config_order", None)
        or getattr(build_doc, "configuration_order", None)
        or getattr(build_doc, "inductone_configuration_order", None)
    )


def _resolve_snapshot_name(build_doc):
    return (
        getattr(build_doc, "selected_snapshot", None)
        or getattr(build_doc, "latest_snapshot", None)
        or getattr(build_doc, "configured_snapshot", None)
        or getattr(build_doc, "configured_bom_snapshot", None)
    )


def _resolve_bom_export_package_name(build, configuration_order):
    candidates = [
        getattr(build, "latest_bom_export_package", None),
        getattr(build, "bom_export_package", None),
        getattr(build, "export_package", None),
        getattr(configuration_order, "bom_export_package", None),
    ]

    for c in candidates:
        if c:
            return c

    for row in (getattr(configuration_order, "documents", None) or []):
        title = (getattr(row, "doc_title", "") or "")
        note = (getattr(row, "small_text_vtsj", "") or "")
        for text in (title, note):
            marker = "BOM Export Package:"
            if marker in text:
                return text.split(marker, 1)[1].split("|")[0].strip()

    return None


def _resolve_flat_bom_file_url(build, configuration_order):
    direct = (
        getattr(build, "flat_bom_file", None)
        or getattr(configuration_order, "flat_bom_file", None)
        or getattr(configuration_order, "configured_flat_bom_file", None)
    )
    if direct:
        return direct

    for row in (getattr(configuration_order, "documents", None) or []):
        title = (getattr(row, "doc_title", "") or "").lower()
        if "flat bom" in title or "configured bom csv" in title:
            return getattr(row, "file_url", None) or getattr(row, "file", None)

    return None


def _get_configured_rows_for_build(build_doc):
    top_bom = getattr(build_doc, "top_bom", None) or getattr(build_doc, "bom", None)
    snapshot_name = _resolve_snapshot_name(build_doc)
    co_name = _resolve_configuration_order_name(build_doc)

    if not top_bom:
        frappe.throw("Top BOM is required to derive configured rows for builder release.")
    if not snapshot_name:
        frappe.throw("Selected/Latest Snapshot is required to derive configured rows for builder release.")
    if not co_name:
        frappe.throw("Latest Configuration Order is required to derive configured rows for builder release.")

    temp = frappe._dict({
        "inductone_build": build_doc.name,
        "configured_snapshot": snapshot_name,
        "bom": top_bom,
        "explosion_mode": getattr(build_doc, "explosion_mode", None) or "Follow Explicit Child BOM Links",
        "max_depth": getattr(build_doc, "max_depth", None),
        "include_qty": getattr(build_doc, "include_qty", 1),
        "configuration_order": co_name,
    })
    return build_configured_rows(temp)


def _derive_required_serial_capture_rows(configured_rows):
    """
    Retained only for later validation/comparison workflows.
    The builder-facing submission artifact is now the workbook, not this generic row set.
    """
    item_meta = {}
    out = []

    for row in configured_rows:
        item_code = row.get("item_code")
        if not item_code:
            continue

        if item_code not in item_meta:
            item_meta[item_code] = frappe.db.get_value(
                "Item",
                item_code,
                ["item_name", "has_serial_no", "serial_no_series", "description"],
                as_dict=True,
            ) or {}

        meta = item_meta[item_code]
        if int(meta.get("has_serial_no") or 0) != 1:
            continue

        out.append({
            "item_code": item_code,
            "item_name": row.get("item_name") or meta.get("item_name") or "",
            "bom_level": row.get("bom_level"),
            "node_type": row.get("node_type"),
            "expected_qty": row.get("qty") or 1,
            "bom_used": row.get("bom_used") or "",
            "parent_item_code": (row.get("ancestor_item_codes") or [""])[-1] if (row.get("ancestor_item_codes") or []) else "",
            "serial_no_series": meta.get("serial_no_series") or "",
            "description": row.get("description") or meta.get("description") or "",
            "vendor_serial_no": "",
            "por_serial_no": "",
            "installed_serial_no": "",
            "installed": 1,
            "notes": "",
        })

    return out


def _required_serial_capture_csv_bytes(rows):
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([
        "item_code", "item_name", "bom_level", "node_type", "expected_qty", "bom_used",
        "parent_item_code", "serial_no_series", "description", "vendor_serial_no",
        "por_serial_no", "installed_serial_no", "installed", "notes"
    ])
    for r in rows:
        w.writerow([
            r.get("item_code") or "",
            r.get("item_name") or "",
            r.get("bom_level") or "",
            r.get("node_type") or "",
            r.get("expected_qty") or "",
            r.get("bom_used") or "",
            r.get("parent_item_code") or "",
            r.get("serial_no_series") or "",
            r.get("description") or "",
            r.get("vendor_serial_no") or "",
            r.get("por_serial_no") or "",
            r.get("installed_serial_no") or "",
            r.get("installed") or "",
            r.get("notes") or "",
        ])
    return out.getvalue().encode("utf-8")


def _as_built_serial_capture_csv_bytes(rows):
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow([
        "item_code", "item_name", "bom_level", "expected_qty", "serial_no", "vendor_serial_no",
        "por_serial_no", "parent_item_code", "installed", "notes"
    ])
    for r in rows:
        w.writerow([
            r.get("item_code") or "",
            r.get("item_name") or "",
            r.get("bom_level") or "",
            r.get("expected_qty") or "",
            r.get("serial_no") or r.get("installed_serial_no") or "",
            r.get("vendor_serial_no") or "",
            r.get("por_serial_no") or "",
            r.get("parent_item_code") or "",
            r.get("installed") if r.get("installed") is not None else 1,
            r.get("notes") or "",
        ])
    return out.getvalue().encode("utf-8")


def _sync_required_serial_capture_table(build_doc, rows):
    fieldname = _find_first_child_table_field(build_doc, [
        "required_serial_capture",
        "required_serials",
        "builder_required_serial_capture",
        "serial_capture_requirements",
    ])
    if not fieldname:
        return

    build_doc.set(fieldname, [])
    for r in rows:
        build_doc.append(fieldname, {
            "item_code": r.get("item_code"),
            "item_name": r.get("item_name"),
            "bom_level": r.get("bom_level"),
            "node_type": r.get("node_type"),
            "expected_qty": r.get("expected_qty"),
            "qty_required": r.get("expected_qty"),
            "bom_used": r.get("bom_used"),
            "parent_item_code": r.get("parent_item_code"),
            "serial_no_series": r.get("serial_no_series"),
            "description": r.get("description"),
            "installed": 1,
            "status": "Pending",
        })
    build_doc.save(ignore_permissions=True)


def _sync_as_built_serial_capture_table(build_doc, rows):
    fieldname = _find_first_child_table_field(build_doc, [
        "as_built_serial_capture",
        "as_built_serials",
        "actual_serial_capture",
        "builder_serial_submission",
    ])
    if not fieldname:
        return

    build_doc.set(fieldname, [])
    for r in rows:
        build_doc.append(fieldname, {
            "item_code": r.get("item_code"),
            "item_name": r.get("item_name"),
            "bom_level": r.get("bom_level"),
            "expected_qty": r.get("expected_qty"),
            "serial_no": r.get("serial_no") or r.get("installed_serial_no"),
            "vendor_serial_no": r.get("vendor_serial_no"),
            "por_serial_no": r.get("por_serial_no"),
            "parent_item_code": r.get("parent_item_code"),
            "installed": r.get("installed") if r.get("installed") is not None else 1,
            "notes": r.get("notes"),
            "status": "Submitted",
        })
    build_doc.save(ignore_permissions=True)


def _sync_bom_export_document_index(configuration_order_name, top_bom, package_doc):
    if not package_doc:
        return

    row_title = f"Configured BOM Export Package - {package_doc.name}"
    row_note = f"BOM Export Package: {package_doc.name} | Status: {package_doc.status or 'Draft'}"

    _append_or_update_document_index_row(
        parent_doctype="InductOne Configuration Order",
        parent_name=configuration_order_name,
        title=row_title,
        file_url=package_doc.output_zip or "",
        source_type="BOM",
        source_name=top_bom or "",
        sort_order=300,
        note=row_note,
    )


def _sync_flat_bom_document_index(configuration_order_name, build_name, flat_bom_file_url):
    if not flat_bom_file_url:
        return

    doc = frappe.get_doc("InductOne Configuration Order", configuration_order_name)
    existing_title = None

    for row in (getattr(doc, "documents", None) or []):
        title = (getattr(row, "doc_title", "") or "")
        if "flat bom" in title.lower() or "configured bom csv" in title.lower():
            existing_title = title
            break

    title_to_use = existing_title or f"{configuration_order_name} Flat BOM CSV"

    _append_or_update_document_index_row(
        parent_doctype="InductOne Configuration Order",
        parent_name=configuration_order_name,
        title=title_to_use,
        file_url=flat_bom_file_url,
        source_type="MANUAL",
        source_name=configuration_order_name,
        sort_order=900,
        note="Auto-generated rolled-up flat BOM CSV.",
    )


def _sync_builder_release_document_index(configuration_order_name, build_name, manifest_url, package_name=None, flat_bom_file_url=None):
    note = f"Builder release manifest for build {build_name}"
    if package_name:
        note += f" | BOM Export Package: {package_name}"
    if flat_bom_file_url:
        note += f" | Flat BOM: {flat_bom_file_url}"

    _append_or_update_document_index_row(
        parent_doctype="InductOne Configuration Order",
        parent_name=configuration_order_name,
        title=f"Builder Release Manifest - {build_name}",
        file_url=manifest_url,
        source_type="MANUAL",
        source_name=build_name,
        sort_order=310,
        note=note,
    )


def _append_or_update_document_index_row(parent_doctype, parent_name, title, file_url, source_type="OTHER", source_name="", sort_order=300, note=""):
    doc = frappe.get_doc(parent_doctype, parent_name)
    existing = None

    for row in (getattr(doc, "documents", None) or []):
        if (getattr(row, "doc_title", "") or "").strip() == title:
            existing = row
            break

    payload = {
        "source_type": source_type,
        "source_name": source_name,
        "doc_type": "OTHER",
        "doc_title": title,
        "file": file_url or "",
        "file_url": file_url or "",
        "required": "YES",
        "sort_order": sort_order,
        "small_text_vtsj": note,
    }

    if existing:
        for k, v in payload.items():
            if hasattr(existing, k):
                setattr(existing, k, v)
    else:
        doc.append("documents", payload)

    doc.save(ignore_permissions=True)


def _find_first_child_table_field(doc, candidates):
    meta = getattr(doc, "meta", None)
    if not meta:
        return None

    for name in candidates:
        try:
            field = meta.get_field(name)
        except Exception:
            field = None
        if field and field.fieldtype == "Table":
            return name

    return None


def _set_if_present(doc, fieldnames, value):
    for fname in fieldnames:
        try:
            if hasattr(doc, fname):
                setattr(doc, fname, value)
                return True
            field = doc.meta.get_field(fname)
            if field:
                doc.set(fname, value)
                return True
        except Exception:
            continue
    return False