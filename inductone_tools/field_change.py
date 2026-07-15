"""Field Change Order (FCO) server-side behavior.

The FCO DocTypes are custom DocTypes managed through fixtures. Frappe Cloud does
not load filesystem controllers for custom DocTypes, so hooks.py routes
validation here and Desk buttons/client scripts can call the whitelisted methods
by dotted path.
"""

from __future__ import annotations

import json
from typing import Iterable

import frappe
from frappe import _
from frappe.utils import getdate, now_datetime
from frappe.utils.file_manager import get_file_path


FCO_ROLES = {
    "Operations Manager",
    "InductOne Manager",
    "InductOne Process Architect",
    "System Manager",
}

LOCKED_STATUS = "Locked"


def _require_fco_role() -> None:
    roles = set(frappe.get_roles(frappe.session.user))
    if not (roles & FCO_ROLES):
        frappe.throw(
            _("This action requires an Operations or InductOne process role."),
            frappe.PermissionError,
        )


def validate_field_change_request(doc, method=None) -> None:
    """Validate tracked-but-changeable Instance assignment.

    Backfilled assignments can be low-confidence, so reassignment must remain
    possible. The control is not immutability; it is a required reason plus
    DocType `track_changes`, which preserves the Version trail.
    """

    _sync_location_display_fields(doc, fallback_to_machine_identifier=True)

    if doc.is_new():
        return

    old_instance = frappe.db.get_value(doc.doctype, doc.name, "instance")
    new_instance = doc.get("instance")
    if old_instance != new_instance and not (doc.get("assignment_change_reason") or "").strip():
        frappe.throw(
            _(
                "Changing the assigned Instance requires an Assignment Change Reason. "
                "This preserves the FCO assignment audit trail."
            )
        )

    if doc.get("assignment_reviewed") and not doc.get("assignment_reviewed_at"):
        doc.assignment_reviewed_at = now_datetime()
        doc.assignment_reviewed_by = frappe.session.user


def validate_field_change(doc, method=None) -> None:
    """Keep accepted Field Change records immutable."""

    _sync_location_display_fields(doc, fallback_to_machine_identifier=False)

    if doc.is_new():
        return

    old_status = frappe.db.get_value(doc.doctype, doc.name, "status")
    if old_status == LOCKED_STATUS:
        frappe.throw(_("Locked Field Change records are immutable."))


def _sync_location_display_fields(doc, *, fallback_to_machine_identifier: bool) -> None:
    """Denormalize Instance customer/location for readable FCO list views.

    These fields are display aids. The Instance remains the canonical link to
    customer and location, but operators need the list view to read as an
    as-installed/as-maintained ledger instead of a bare serial-number table.
    """

    if not hasattr(doc, "meta"):
        return

    has_location_label = doc.meta.has_field("location_label")
    has_customer = doc.meta.has_field("customer")
    if not has_location_label and not has_customer:
        return

    instance_name = doc.get("instance")
    if instance_name:
        values = frappe.db.get_value(
            "InductOne Instance",
            instance_name,
            ["deployment_site", "customer"],
            as_dict=True,
        ) or {}
        if has_location_label:
            doc.location_label = values.get("deployment_site")
        if has_customer:
            doc.customer = values.get("customer")
        return

    if has_location_label:
        doc.location_label = doc.get("machine_identifier") if fallback_to_machine_identifier else None
    if has_customer:
        doc.customer = None


def refresh_display_labels() -> dict:
    """Backfill denormalized FCO list-view labels after fixture migration.

    This is intentionally not whitelisted. It is called from after_migrate,
    after the DocType fixture has created the ``location_label`` and
    ``customer`` columns.
    """

    result = {"requests": 0, "field_changes": 0, "skipped": []}

    if frappe.db.exists("DocType", "InductOne Field Change Request") and _has_columns(
        "InductOne Field Change Request", ["location_label", "customer"]
    ):
        for row in frappe.get_all(
            "InductOne Field Change Request",
            fields=["name", "instance", "machine_identifier"],
        ):
            frappe.db.set_value(
                "InductOne Field Change Request",
                row.name,
                _display_values(row.instance, row.machine_identifier),
                update_modified=False,
            )
            result["requests"] += 1
    else:
        result["skipped"].append("InductOne Field Change Request")

    if frappe.db.exists("DocType", "InductOne Field Change") and _has_columns(
        "InductOne Field Change", ["location_label", "customer"]
    ):
        for row in frappe.get_all("InductOne Field Change", fields=["name", "instance"]):
            frappe.db.set_value(
                "InductOne Field Change",
                row.name,
                _display_values(row.instance, None),
                update_modified=False,
            )
            result["field_changes"] += 1
    else:
        result["skipped"].append("InductOne Field Change")

    return result


def _has_columns(doctype: str, columns: list[str]) -> bool:
    return all(frappe.db.has_column(doctype, column) for column in columns)


def _display_values(instance: str | None, fallback_location: str | None) -> dict:
    if instance:
        values = frappe.db.get_value(
            "InductOne Instance",
            instance,
            ["deployment_site", "customer"],
            as_dict=True,
        ) or {}
        return {
            "location_label": values.get("deployment_site"),
            "customer": values.get("customer"),
        }
    return {"location_label": fallback_location, "customer": None}


@frappe.whitelist()
def accept_field_change(field_change: str) -> dict:
    """Accept a per-Instance Field Change and update the Instance state.

    This is intentionally permission-respecting. It gates by role first and then
    uses ordinary document saves; no `ignore_permissions=True` appears in this
    whitelisted state-changing path.
    """

    _require_fco_role()
    if not field_change:
        frappe.throw(_("Field Change name is required."))

    doc = frappe.get_doc("InductOne Field Change", field_change)
    if doc.status == LOCKED_STATUS:
        return {"ok": True, "field_change": doc.name, "status": doc.status}

    if not doc.instance:
        frappe.throw(_("Field Change must be linked to an InductOne Instance before acceptance."))
    if not doc.change_summary:
        frappe.throw(_("Field Change must include a change summary before acceptance."))

    doc.status = LOCKED_STATUS
    doc.accepted_by = frappe.session.user
    doc.accepted_at = now_datetime()
    doc.as_maintained_updated = 1
    if not doc.implemented_date:
        doc.implemented_date = frappe.utils.today()
    doc.save()

    _apply_serial_changes_to_instance(doc)
    _sync_instance_field_change_rollup(doc.instance)

    frappe.db.commit()
    return {
        "ok": True,
        "field_change": doc.name,
        "instance": doc.instance,
        "status": doc.status,
    }


def _apply_serial_changes_to_instance(field_change_doc) -> None:
    if not field_change_doc.component_serial_changes:
        return

    instance = frappe.get_doc("InductOne Instance", field_change_doc.instance)
    serials = list(instance.get("component_serials") or [])

    for change in field_change_doc.component_serial_changes:
        component = (change.component or "").strip()
        action = change.action
        old_serial = (change.old_serial or "").strip()
        new_serial = (change.new_serial or "").strip()

        if action in {"Remove", "Replace"} and old_serial:
            for row in list(serials):
                if row.component_label == component and row.serial_number == old_serial:
                    instance.remove(row)
                    serials.remove(row)
                    break

        if action in {"Add", "Replace"} and new_serial:
            exists = any(
                row.component_label == component and row.serial_number == new_serial
                for row in instance.get("component_serials")
            )
            if not exists:
                instance.append(
                    "component_serials",
                    {
                        "component_label": component,
                        "serial_number": new_serial,
                    },
                )

    instance.save()


def _sync_instance_field_change_rollup(instance_name: str) -> None:
    latest = frappe.db.sql(
        """
        SELECT COUNT(*) AS count, MAX(implemented_date) AS latest_date
        FROM `tabInductOne Field Change`
        WHERE instance = %s AND status = 'Locked'
        """,
        instance_name,
        as_dict=True,
    )[0]

    values = {
        "field_change_count": latest.count or 0,
        "latest_field_change_date": latest.latest_date,
    }
    frappe.db.set_value("InductOne Instance", instance_name, values, update_modified=False)


@frappe.whitelist()
def render_fco_register() -> list[dict]:
    """Return the SUP-FCO-R01 v2.0 register projection as JSON rows."""

    frappe.only_for(tuple(FCO_ROLES | {"Operations Viewer"}))
    return frappe.db.sql(
        """
        SELECT
          r.name AS fco_number,
          r.date_raised,
          r.requester,
          r.intake_ref,
          r.customer_project,
          COALESCE(r.instance, r.machine_identifier) AS serial_or_location,
          r.title AS change_summary,
          r.triage_outcome,
          r.reference,
          IF(r.safety_regulatory, 'Y', 'N') AS safety_regulatory,
          r.disposition,
          r.disposition_date,
          MAX(fc.implemented_date) AS implemented_date,
          IF(MAX(IFNULL(fc.as_maintained_updated, 0)) = 1, 'Y', 'N') AS as_maintained_updated,
          MAX(fc.post_change_test) AS post_change_test,
          r.status,
          IF(r.status='Closed', r.modified, NULL) AS closed_date,
          r.notes
        FROM `tabInductOne Field Change Request` r
        LEFT JOIN `tabInductOne Field Change` fc ON fc.source_request = r.name
        GROUP BY r.name
        ORDER BY r.date_raised, r.name
        """,
        as_dict=True,
    )


@frappe.whitelist()
def import_jotform_export(file_url: str) -> dict:
    """Import a JotForm FCO export workbook into Request records.

    This is the reusable, permission-respecting operator path for post-launch
    intake. The historical seed backfill remains in scripts/ because it also
    applies owner-reviewed assignment guesses and one-time seeded Field Changes.
    """

    _require_fco_role()
    if not file_url:
        frappe.throw(_("A JotForm XLSX File URL is required."))

    rows = _jotform_workbook_rows(file_url)
    created: list[str] = []
    existing: list[str] = []
    skipped: list[dict] = []

    for row in rows:
        intake_ref = _intake_ref_from_jotform_row(row)
        if not intake_ref:
            skipped.append({"row": row, "reason": "missing Change No / date"})
            continue

        if frappe.db.exists("InductOne Field Change Request", intake_ref):
            existing.append(intake_ref)
            continue

        status, disposition = _disposition_from_flow(row.get("Flow Status"))
        doc = frappe.get_doc(
            {
                "doctype": "InductOne Field Change Request",
                "name": intake_ref,
                "naming_series": "FCO-.YYYY.-.###",
                "intake_ref": intake_ref,
                "status": status,
                "date_raised": _parse_jotform_date(row.get("Date") or row.get("Submission Date")),
                "requester": row.get("Change Requestor"),
                "requester_department": row.get("Department"),
                "requester_role": row.get("Role"),
                "intake_source": "JotForm Import",
                "title": row.get("Title of Field Change Order") or intake_ref,
                "description": row.get("Description of proposed change in plain English:"),
                "reason": row.get("Benefit/Reason for change:"),
                "customer_project": row.get("What site(s), project(s), or customer does this apply to?"),
                "one_time_or_repeated": row.get(
                    "Do you expect this to be a one-time fix, or repeated across sites?"
                ),
                "est_downtime_h": row.get("Estimated Customer Downtime hours"),
                "est_labor_h": row.get("Estimated Labor Hours"),
                "parts_cost": row.get("Part(s) Cost ($)"),
                "implementer": row.get("Who can implement this field change?"),
                "tools_docs": row.get("Tools / Documents Needed"),
                "ticket_link": row.get("Relevant issue tracker link or ticket:"),
                "triage_outcome": _triage_from_jotform_row(row),
                "reference": row.get("ECR Link") or row.get("Epic link"),
                "safety_regulatory": 0,
                "disposition": disposition,
                "disposition_date": _parse_jotform_date(row.get("Submittal Date")),
                "disposition_by": row.get("Approver Email"),
                "assignment_confidence": "Backfill-guess",
                "assignment_reviewed": 0,
                "notes": row.get("Notes"),
            }
        )
        doc.insert()
        created.append(intake_ref)

    frappe.db.commit()
    return {
        "ok": True,
        "file_url": file_url,
        "rows_read": len(rows),
        "created": created,
        "existing": existing,
        "skipped": skipped,
    }


def request_rows_from_json(payload: str | Iterable[dict]) -> list[dict]:
    """Small helper used by tests/importers to normalize row payloads."""

    if isinstance(payload, str):
        return json.loads(payload)
    return list(payload)


def _jotform_workbook_rows(file_url: str) -> list[dict]:
    from openpyxl import load_workbook

    workbook_path = get_file_path(file_url)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    rows = []
    for values in iterator:
        row = {
            headers[index]: value
            for index, value in enumerate(values)
            if index < len(headers) and headers[index]
        }
        if any(value not in (None, "") for value in row.values()):
            rows.append(row)
    return rows


def _intake_ref_from_jotform_row(row: dict) -> str | None:
    change_no = row.get("Change No")
    if change_no in (None, ""):
        return None

    try:
        suffix = f"{int(float(change_no)):03d}"
    except (TypeError, ValueError):
        suffix = str(change_no).strip().zfill(3)

    date_value = _parse_jotform_date(row.get("Date") or row.get("Submission Date"))
    if not date_value:
        return None
    return f"FCO-{date_value.year}-{suffix}"


def _parse_jotform_date(value):
    if not value:
        return None
    try:
        return getdate(value)
    except Exception:
        return getdate()


def _disposition_from_flow(flow_status: str | None) -> tuple[str, str]:
    value = (flow_status or "").strip().lower()
    if value == "complete":
        return "Closed", "Approved"
    if value == "denied":
        return "Rejected", "Rejected"
    if value in {"canceled", "cancelled"}:
        return "Cancelled", "Cancelled"
    return "Triaged", "Pending"


def _triage_from_jotform_row(row: dict) -> str:
    if row.get("ECR Link"):
        return "ECR"
    return "Field Change"
