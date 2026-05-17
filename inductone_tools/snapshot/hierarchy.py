"""
Configured BOM Snapshot Hierarchy — populator and workbook generator.

Two whitelisted entry points:

  populate_snapshot_hierarchy(snapshot_name):
      Called from the client snapshot-generation flow right after the
      Configured BOM Snapshot is inserted. Reads the snapshot, runs the
      existing tree resolver (bom_export.build_configured_rows), and
      writes one Configured BOM Snapshot Hierarchy row per resolved
      node. Frozen forever after.

  generate_hierarchy_workbook(snapshot_name):
      Called immediately after populate_snapshot_hierarchy. Reads the
      hierarchy rows, renders an XLSX matching the live BOM Explorer
      report visual style, attaches it to the Snapshot, registers it
      in the linked CO's Document Index.

Design notes:
  - Path 1 architecture: snapshot creation is client-driven. These
    server functions are called explicitly by the client AFTER the
    snapshot insert succeeds. No doc_events, no after_insert magic.
  - We use bom_export.build_configured_rows directly (Path B2 from the
    discussion): a stub BOM Export Package doc is constructed in memory
    and passed in. The export-package code path is unchanged.
  - Item Name, Description, UOM, Item Group are copied from live Item
    records at hierarchy-population time. Once written, they are frozen
    on the hierarchy child rows and never re-read. This is the whole
    point — the snapshot is immune to later edits.
  - Excel outline grouping is applied per bom_level so builders get
    native collapse/expand. Indentation in the Item Code column uses
    leading nbsp characters so visual hierarchy survives sort/filter.
"""

import io

import frappe
from frappe import _
from frappe.utils import now, now_datetime
from frappe.utils.file_manager import save_file

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from inductone_tools.bom_export import build_configured_rows


# ============================================================
#  Public entry points
# ============================================================

@frappe.whitelist()
def populate_snapshot_hierarchy(snapshot_name):
    """
    Materialize the hierarchical tree onto the Configured BOM Snapshot.

    Called from the client snapshot-generation flow after the snapshot is
    inserted. Writes one Configured BOM Snapshot Hierarchy row per node
    in the resolved tree. Idempotent: existing hierarchy rows are wiped
    and replaced if called again on the same snapshot.

    Returns a dict with row counts and any warnings.
    """
    if not snapshot_name:
        frappe.throw(_("snapshot_name is required."))

    snap = frappe.get_doc("Configured BOM Snapshot", snapshot_name)

    if not snap.inductone_build:
        frappe.throw(_(
            "Snapshot {0} has no linked InductOne Build; "
            "hierarchy population requires it to resolve options."
        ).format(snapshot_name))

    if not snap.top_bom:
        frappe.throw(_(
            "Snapshot {0} has no top_bom; hierarchy cannot be resolved."
        ).format(snapshot_name))

    # Wipe any existing hierarchy rows. Idempotent re-runs replace
    # rather than append.
    _clear_snapshot_hierarchy(snapshot_name)

    # Build a stub BOM Export Package doc so we can call
    # build_configured_rows without modifying it.
    stub_package = _build_stub_export_package(snap)

    # This is where the work happens. The function resolves baseline
    # tree + structural effects + selected options into a final
    # filtered list of nodes.
    resolved_rows = build_configured_rows(stub_package)

    # Walk resolved rows, attach node_id and parent_node_id, write to
    # the hierarchy child table.
    hierarchy_rows = _assign_node_ids_and_parents(resolved_rows)

    # Enrich each row with frozen copies of Item metadata.
    _enrich_with_item_metadata(hierarchy_rows)

    # Bulk insert.
    _insert_hierarchy_rows(snapshot_name, hierarchy_rows)

    frappe.db.commit()

    return {
        "ok": True,
        "snapshot_name": snapshot_name,
        "hierarchy_rows": len(hierarchy_rows),
        "assemblies": sum(1 for r in hierarchy_rows if r["node_type"] == "Assembly"),
        "leaves": sum(1 for r in hierarchy_rows if r["node_type"] == "Leaf"),
    }


@frappe.whitelist()
def generate_hierarchy_workbook(snapshot_name):
    """
    Generate the indented BOM XLSX from the snapshot's hierarchy table,
    attach it to the snapshot, and register it in the linked CO's
    Document Index.

    Returns the file_url of the saved workbook.
    """
    if not snapshot_name:
        frappe.throw(_("snapshot_name is required."))

    snap = frappe.get_doc("Configured BOM Snapshot", snapshot_name)

    if not (snap.hierarchy or []):
        frappe.throw(_(
            "Snapshot {0} has no hierarchy rows. Run populate_snapshot_hierarchy "
            "first, or regenerate the snapshot."
        ).format(snapshot_name))

    workbook_bytes = _render_hierarchy_workbook(snap)

    fname = "{0}_Configured_BOM_Hierarchy_{1}.xlsx".format(
        snapshot_name,
        now_datetime().strftime("%Y%m%d_%H%M%S")
    )

    saved = save_file(
        fname=fname,
        content=workbook_bytes,
        dt="Configured BOM Snapshot",
        dn=snapshot_name,
        is_private=1,
    )

    # Register in CO Document Index if there's a linked CO.
    co_name = _resolve_co_for_snapshot(snapshot_name)
    if co_name:
        _register_in_co_document_index(co_name, snapshot_name, saved.file_url)

    frappe.db.commit()

    return {
        "ok": True,
        "snapshot_name": snapshot_name,
        "file_url": saved.file_url,
        "configuration_order": co_name,
    }

@frappe.whitelist()
def sync_hierarchy_workbook_to_configuration_order(snapshot_name, co_name=None):
    """
    Register the latest existing Configured BOM Hierarchy workbook
    attached to a Configured BOM Snapshot into the linked Configuration
    Order Document Index.

    This does not generate a new workbook. It only reconciles an
    already-attached snapshot artifact into the CO documents table.
    """
    if not snapshot_name:
        frappe.throw(_("snapshot_name is required."))

    # Validate snapshot exists.
    frappe.get_doc("Configured BOM Snapshot", snapshot_name)

    if not co_name:
        co_name = _resolve_co_for_snapshot(snapshot_name)

    if not co_name:
        return {
            "ok": False,
            "snapshot_name": snapshot_name,
            "configuration_order": None,
            "file_url": None,
            "message": "No Configuration Order found for this snapshot.",
        }

    files = frappe.get_all(
        "File",
        filters={
            "attached_to_doctype": "Configured BOM Snapshot",
            "attached_to_name": snapshot_name,
        },
        fields=["name", "file_name", "file_url", "creation"],
        order_by="creation desc",
        limit_page_length=50,
    )

    hierarchy_file = None

    for f in files:
        file_name = (f.get("file_name") or "").lower()
        file_url = (f.get("file_url") or "").lower()

        if (
            "configured_bom_hierarchy" in file_name
            or "configured bom hierarchy" in file_name
            or "configured_bom_hierarchy" in file_url
            or "configured bom hierarchy" in file_url
            or "_configured_bom_hierarchy_" in file_name
            or "_configured_bom_hierarchy_" in file_url
        ):
            hierarchy_file = f
            break

    if not hierarchy_file:
        return {
            "ok": False,
            "snapshot_name": snapshot_name,
            "configuration_order": co_name,
            "file_url": None,
            "message": "No Configured BOM Hierarchy workbook attachment found on snapshot.",
        }

    file_url = hierarchy_file.get("file_url")
    if not file_url:
        return {
            "ok": False,
            "snapshot_name": snapshot_name,
            "configuration_order": co_name,
            "file_url": None,
            "message": "Hierarchy workbook File record has no file_url.",
        }

    _register_in_co_document_index(co_name, snapshot_name, file_url)

    frappe.db.commit()

    return {
        "ok": True,
        "snapshot_name": snapshot_name,
        "configuration_order": co_name,
        "file_url": file_url,
        "file_name": hierarchy_file.get("file_name"),
        "message": "Configured BOM Hierarchy workbook synced to Configuration Order.",
    }

# ============================================================
#  Stub package construction — Path B2
# ============================================================

def _build_stub_export_package(snap):
    """
    Construct a minimal BOM Export Package doc in memory with just
    enough fields for build_configured_rows to do its work.

    This is the Path B2 approach: rather than refactor
    build_configured_rows to accept a generic context, we hand it the
    same shape of object it already expects. No risk of behavior drift.

    Not inserted, not saved. Lives only in memory for the duration of
    this call.
    """
    stub = frappe.get_doc({
    "doctype": "BOM Export Package",
    "source_mode": "Configured Build",
    "bom": snap.top_bom,
    "inductone_build": snap.inductone_build,
    "configured_snapshot": snap.name,
    # build_configured_rows reads explosion_mode and max_depth.
    # Use the same defaults the real release flow uses.
    "explosion_mode": "Follow Explicit Child BOM Links",
    "max_depth": 0,
    "include_qty": 1,
    })

    # Hierarchy-specific resolver behavior:
    #
    # 1. Preserve duplicate BOM row occurrences so the frozen workbook
    #    matches the live BOM report structure.
    # 2. Apply configured snapshot quantities where safe, e.g. the
    #    25 ft hose override, without stamping rolled-up flat quantities
    #    onto repeated branch occurrences.
    stub.preserve_duplicate_occurrences = 1
    stub.apply_snapshot_quantities = 1

    return stub


# ============================================================
#  Hierarchy row construction
# ============================================================

def _assign_node_ids_and_parents(resolved_rows):
    """
    Build hierarchy rows with stable node_id / parent_node_id linkage
    using each row's ancestor path, not incoming row order.

    Why this exists:
      build_configured_rows() is optimized for configured export output
      and may return rows sorted/grouped by bom_level, item_code, etc.
      That order is not a valid tree walk. The previous implementation
      used a stack and assumed rows arrived depth-first, which caused
      level-2+ rows to attach to the most recent prior row at the previous
      level instead of their real parent.

    Canonical rule:
      Each resolved row carries:
        - bom_level
        - item_code
        - ancestor_item_codes

      For a row:
        full_path   = ancestor_item_codes + [item_code]
        parent_path = ancestor_item_codes

      The parent is the row whose full_path equals parent_path.

    Notes:
      - Item codes can repeat in different branches, so parent lookup must
        be path-based, not item-code-based.
      - Duplicate item rows under the exact same parent path are allowed;
        they receive occurrence suffixes internally so node IDs remain
        unique while the visible item data remains unchanged.
      - Output order is rebuilt into a depth-first tree order so the
        workbook's indentation and Excel outline controls behave like a
        real hierarchy.
    """
    if not resolved_rows:
        return []

    normalized = []

    # First pass: normalize rows and create durable path keys.
    #
    # A raw path of only item codes is sufficient for parent lookup in the
    # normal case. However, duplicate item rows can exist under the same
    # parent. To avoid overwriting path_index entries, we track occurrences
    # of the same raw path and append an internal occurrence number.
    #
    # Parent lookup intentionally uses the unsuffixed parent raw path and
    # resolves to the first matching parent node for that path, which is
    # correct for normal BOM assembly parentage.
    raw_path_counts = {}

    for original_index, row in enumerate(resolved_rows):
        item_code = row.get("item_code") or ""
        bom_level = int(row.get("bom_level") or 0)
        ancestor_items = list(row.get("ancestor_item_codes") or [])

        full_raw_path = tuple(ancestor_items + [item_code])
        parent_raw_path = tuple(ancestor_items)

        occurrence = raw_path_counts.get(full_raw_path, 0) + 1
        raw_path_counts[full_raw_path] = occurrence

        normalized.append({
            "original_index": original_index,
            "row": row,
            "item_code": item_code,
            "bom_level": bom_level,
            "ancestor_items": ancestor_items,
            "full_raw_path": full_raw_path,
            "parent_raw_path": parent_raw_path,
            "occurrence": occurrence,
            "children": [],
            "node_id": None,
            "parent_node_id": "",
        })

    # Index the first node seen for each raw path. This is used for parent
    # lookup. Because a valid BOM parent assembly should normally be unique
    # at a given path, first-node lookup is appropriate and stable.
    first_node_by_raw_path = {}
    all_nodes_by_raw_path = {}

    for node in normalized:
        full_raw_path = node["full_raw_path"]
        all_nodes_by_raw_path.setdefault(full_raw_path, []).append(node)
        if full_raw_path not in first_node_by_raw_path:
            first_node_by_raw_path[full_raw_path] = node

    roots = []

    # Second pass: link each node to its parent using ancestor path.
    for node in normalized:
        parent_raw_path = node["parent_raw_path"]

        if not parent_raw_path:
            # Immediate child of the synthetic workbook root.
            roots.append(node)
            continue

        parent_node = first_node_by_raw_path.get(parent_raw_path)

        if parent_node:
            node["parent_node_id"] = "__PENDING__"
            parent_node["children"].append(node)
        else:
            # Defensive fallback:
            # If a configured structural addition has a parent path that is
            # not present in the filtered result, keep it as a root-level row
            # rather than throwing away data or attaching it to a wrong parent.
            #
            # This should be rare. It preserves visibility while avoiding
            # false hierarchy.
            roots.append(node)

    # Sort siblings in a stable, BOM-report-like order.
    #
    # Use original_index as the final tiebreaker so we preserve as much of
    # the upstream resolver's deterministic ordering as possible within a
    # given parent branch.
    def _sibling_sort_key(node):
        row = node["row"]
        return (
            int(row.get("row_order") or 0),
            int(row.get("bom_level") or 0),
            row.get("item_code") or "",
            row.get("bom_used") or "",
            row.get("node_type") or "",
            node["occurrence"],
            node["original_index"],
        )

    def _sort_tree(nodes):
        nodes.sort(key=_sibling_sort_key)
        for n in nodes:
            _sort_tree(n["children"])

    _sort_tree(roots)

    # Third pass: emit in depth-first order and assign node IDs.
    out = []
    counter = [0]

    def _next_id():
        counter[0] += 1
        return "N{0:05d}".format(counter[0])

    def _emit(node, parent_node_id=""):
        row = node["row"]
        node_id = _next_id()
        node["node_id"] = node_id

        bom_level = int(row.get("bom_level") or 0)

        hierarchy_node = {
            "node_id": node_id,
            "parent_node_id": parent_node_id,
            "bom_level": bom_level,
            "item_code": row.get("item_code") or "",
            "item_name": row.get("item_name") or "",
            "description": row.get("description") or "",
            "qty": float(row.get("qty") or 0),
            "uom": row.get("uom") or "",
            "bom_used": row.get("bom_used") or "",
            "node_type": row.get("node_type") or "Leaf",
            "is_leaf": 1 if row.get("is_leaf") else 0,
            # effect_origin and source_option_code are not currently in
            # the resolved row dict from build_configured_rows; we set
            # BASELINE as default and let future work refine this when
            # build_configured_rows is updated to propagate origin.
            "effect_origin": row.get("effect_origin") or "BASELINE",
            "source_option_code": row.get("source_option_code") or "",
            "excluded_by_structural_effect": int(row.get("excluded_by_structural_effect") or 0),
            "item_group": "",  # filled in by _enrich_with_item_metadata
            "source_bom": row.get("source_bom") or "",
            "source_bom_item": row.get("source_bom_item") or "",
            "source_bom_item_idx": int(row.get("source_bom_item_idx") or 0),
            "balloon_numbers": row.get("balloon_numbers") or "",
            "electrical_unit": row.get("electrical_unit") or "",
            "source_electrical_bom_rev": row.get("source_electrical_bom_rev") or "",
        }

        out.append(hierarchy_node)

        for child in node["children"]:
            _emit(child, parent_node_id=node_id)

    for root in roots:
        _emit(root, parent_node_id="")

    return out


def _enrich_with_item_metadata(hierarchy_rows):
    """
    Fill item_group, and backfill item_name/description/uom from the
    Item master if any of them are missing on the hierarchy row.

    These values are frozen onto the hierarchy rows. Once written, they
    never re-read live data — the snapshot is immune to later edits.
    """
    item_codes = sorted({
        r["item_code"] for r in hierarchy_rows
        if r.get("item_code")
    })

    if not item_codes:
        return

    # One bulk query for all item metadata.
    items = frappe.get_all(
        "Item",
        filters={"name": ["in", item_codes]},
        fields=["name", "item_name", "description", "stock_uom", "item_group"],
    )
    item_index = {i["name"]: i for i in items}

    for r in hierarchy_rows:
        ic = r.get("item_code")
        if not ic:
            continue
        meta = item_index.get(ic) or {}
        if not r.get("item_name"):
            r["item_name"] = meta.get("item_name") or ""
        if not r.get("description"):
            r["description"] = meta.get("description") or ""
        if not r.get("uom"):
            r["uom"] = meta.get("stock_uom") or ""
        r["item_group"] = meta.get("item_group") or ""


# ============================================================
#  Persistence
# ============================================================

def _clear_snapshot_hierarchy(snapshot_name):
    """Remove all existing hierarchy rows for this snapshot. Used to
    make populate_snapshot_hierarchy idempotent."""
    frappe.db.delete("Configured BOM Snapshot Hierarchy", {
        "parent": snapshot_name,
        "parenttype": "Configured BOM Snapshot",
        "parentfield": "hierarchy",
    })


def _insert_hierarchy_rows(snapshot_name, hierarchy_rows):
    """
    Direct child-row inserts without saving the parent. Same pattern
    used in bom_export.update_results_and_missing_summary — avoids the
    optimistic-lock pitfall on the snapshot form.
    """
    for idx, row in enumerate(hierarchy_rows, start=1):
        doc = frappe.get_doc({
            "doctype": "Configured BOM Snapshot Hierarchy",
            "parent": snapshot_name,
            "parenttype": "Configured BOM Snapshot",
            "parentfield": "hierarchy",
            "idx": idx,
            "node_id": row["node_id"],
            "parent_node_id": row["parent_node_id"],
            "bom_level": row["bom_level"],
            "item_code": row["item_code"],
            "item_name": row["item_name"],
            "description": row["description"],
            "qty": row["qty"],
            "uom": row["uom"],
            "bom_used": row["bom_used"],
            "node_type": row["node_type"],
            "is_leaf": row["is_leaf"],
            "effect_origin": row["effect_origin"],
            "source_option_code": row["source_option_code"],
            "excluded_by_structural_effect": row["excluded_by_structural_effect"],
            "item_group": row["item_group"],
            "source_bom": row.get("source_bom") or "",
            "source_bom_item": row.get("source_bom_item") or "",
            "source_bom_item_idx": int(row.get("source_bom_item_idx") or 0),
            "balloon_numbers": row.get("balloon_numbers") or "",
            "electrical_unit": row.get("electrical_unit") or "",
            "source_electrical_bom_rev": row.get("source_electrical_bom_rev") or "",
        })
        doc.insert(ignore_permissions=True)


# ============================================================
#  Workbook rendering — match BOM Explorer visual style
# ============================================================

# Column layout — matches the BOM Explorer report screenshot.
COLUMNS = [
    ("Item Code",            "item_code",                  50),
    ("Item Name",            "item_name",                  30),
    ("Balloon #",            "balloon_numbers",            14),
    ("Electrical Unit",      "electrical_unit",            16),
    ("Source Rev",           "source_electrical_bom_rev",  14),
    ("BOM",                  "bom_used",                   28),
    ("Qty",                  "qty",                         8),
    ("UOM",                  "uom",                         8),
    ("BOM Level",            "bom_level",                  10),
    ("Standard Description", "description",                40),
    ("Item Group",           "item_group",                 18),

    # Audit/debug columns. Keep visible for now during validation.
    # Hide later if builders do not need them.
    ("Source BOM",           "source_bom",                 28),
    ("Source BOM Item",      "source_bom_item",            18),
    ("Source BOM Item IDX",  "source_bom_item_idx",        12),
]

INDENT_CHAR = "\u00a0"  # non-breaking space
INDENT_WIDTH = 4        # nbsp chars per level

HEADER_FILL = PatternFill("solid", fgColor="F0F0F0")
HEADER_FONT = Font(bold=True, color="333333", size=11)
ROOT_FILL = PatternFill("solid", fgColor="FAFAFA")
ASSEMBLY_FONT = Font(bold=False, color="000000")
LEAF_FONT = Font(bold=False, color="333333")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="E0E0E0"),
)


def _render_hierarchy_workbook(snap):
    """
    Produce the XLSX bytes matching the live BOM Explorer report
    visual style: indented item codes, Excel outline grouping for
    collapse/expand, frozen header, filterable columns, provenance
    header block.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Configured BOM Hierarchy"

    # ---- Provenance header block (rows 1-7) ----
    _write_provenance_header(ws, snap)

    # ---- Column header row (row 9) ----
    header_row = 9
    for col_idx, (label, _key, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Filterable range — applied after we know how many data rows.

    # ---- Data rows ----
    data_start_row = header_row + 1
    current_row = data_start_row

    # Insert the synthetic root row first — the top BOM itself.
    # The BOM Explorer report shows the root assembly as level 0.
    root_row = _build_root_row(snap)
    _write_data_row(ws, current_row, root_row, level=0)
    current_row += 1

    for h in (snap.hierarchy or []):
        row_data = {
            "item_code": h.item_code or "",
            "item_name": h.item_name or "",
            "balloon_numbers": getattr(h, "balloon_numbers", None) or "",
            "electrical_unit": getattr(h, "electrical_unit", None) or "",
            "source_electrical_bom_rev": getattr(h, "source_electrical_bom_rev", None) or "",
            "bom_used": h.bom_used or "",
            "qty": h.qty,
            "uom": h.uom or "",
            "bom_level": h.bom_level or 0,
            "description": h.description or "",
            "item_group": h.item_group or "",
            "source_bom": getattr(h, "source_bom", None) or "",
            "source_bom_item": getattr(h, "source_bom_item", None) or "",
            "source_bom_item_idx": int(getattr(h, "source_bom_item_idx", 0) or 0),
            "node_type": h.node_type or "Leaf",
        }
        _write_data_row(ws, current_row, row_data, level=int(h.bom_level or 0))
        current_row += 1

    data_end_row = current_row - 1

    # ---- Filter + autosizing tweaks ----
    if data_end_row >= data_start_row:
        last_col_letter = get_column_letter(len(COLUMNS))
        ws.auto_filter.ref = "A{0}:{1}{2}".format(
            header_row, last_col_letter, data_end_row
        )

    # ---- Outline grouping ----
    # Apply outline_level so Excel's group-collapse controls work.
    for excel_row in range(data_start_row, data_end_row + 1):
        # We need the bom_level for THIS row, not the iterator.
        # Walk the data we already wrote: read cell (row, BOM Level col).
        bom_level_col = next(i for i, c in enumerate(COLUMNS, start=1) if c[1] == "bom_level")
        level_val = ws.cell(row=excel_row, column=bom_level_col).value
        try:
            level_int = int(level_val or 0)
        except (TypeError, ValueError):
            level_int = 0
        # Outline level: 0 means the row is always visible; > 0 means it
        # collapses under the row above it at the next-lower level.
        ws.row_dimensions[excel_row].outline_level = level_int

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = False

    # ---- Save to bytes ----
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _write_provenance_header(ws, snap):
    """
    Render rows 1-7: title and key metadata block, matching the visual
    weight of the BOM Explorer's header callout.
    """
    title_cell = ws.cell(row=1, column=1, value="Configured BOM Hierarchy")
    title_cell.font = Font(bold=True, size=16, color="1F2A44")

    # Top BOM in a styled callout — matches the report screenshot's
    # rounded BOM name pill.
    top_bom_cell = ws.cell(row=2, column=1, value=snap.top_bom or "(no top BOM)")
    top_bom_cell.font = Font(bold=True, size=11, color="1F2A44")
    top_bom_cell.fill = PatternFill("solid", fgColor="EEF2F7")
    top_bom_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Metadata block rows 4-7
    meta = [
        ("Snapshot",            snap.name),
        ("InductOne Build",     snap.inductone_build or ""),
        ("Top BOM",             snap.top_bom or ""),
        ("Generated At",        str(snap.generated_at or "")),
        ("Snapshot Rev",        str(snap.snapshot_rev or "")),
        ("Source Watermark",    "Configured BOM Hierarchy export | "
                                "snapshot frozen at generation | "
                                "Plus One Robotics confidential"),
    ]
    for i, (label, value) in enumerate(meta, start=4):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.font = Font(bold=True, color="555555", size=10)
        ws.cell(row=i, column=2, value=value).font = Font(color="333333", size=10)


def _build_root_row(snap):
    """Synthesize the root row for the top BOM, since the resolved
    rows from build_configured_rows are children of the root, not
    inclusive of it."""
    top_item = snap.top_item or ""
    top_item_name = ""
    top_item_desc = ""
    top_item_uom = ""
    top_item_group = ""

    if top_item:
        item_meta = frappe.db.get_value(
            "Item", top_item,
            ["item_name", "description", "stock_uom", "item_group"],
            as_dict=True,
        ) or {}
        top_item_name = item_meta.get("item_name") or ""
        top_item_desc = item_meta.get("description") or ""
        top_item_uom = item_meta.get("stock_uom") or ""
        top_item_group = item_meta.get("item_group") or ""

    return {
        "item_code": top_item,
        "item_name": top_item_name,
        "bom_used": snap.top_bom or "",
        "qty": 1,
        "uom": top_item_uom,
        "bom_level": 0,
        "description": top_item_desc,
        "item_group": top_item_group,
        "node_type": "Assembly",
        "source_bom": "",
        "source_bom_item": "",
        "source_bom_item_idx": 0,
        "balloon_numbers": "",
        "electrical_unit": "",
        "source_electrical_bom_rev": "",
    }


def _write_data_row(ws, excel_row, row_data, level):
    """Write a single data row with appropriate indentation and styling."""
    for col_idx, (_label, key, _width) in enumerate(COLUMNS, start=1):
        value = row_data.get(key, "")

        # Indent the Item Code column by bom_level.
        if key == "item_code" and value:
            indent = INDENT_CHAR * (INDENT_WIDTH * level)
            cell_value = "{0}{1}".format(indent, value)
        elif key == "qty":
            # Render qty as float; preserve int display when whole.
            try:
                q = float(value or 0)
                cell_value = q
            except (TypeError, ValueError):
                cell_value = value
        else:
            cell_value = value

        cell = ws.cell(row=excel_row, column=col_idx, value=cell_value)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=False)

        # Slight visual differentiation: assemblies bold, leaves plain.
        if row_data.get("node_type") == "Assembly":
            cell.font = Font(bold=False, color="000000", size=10)
        else:
            cell.font = Font(bold=False, color="333333", size=10)

        if key == "qty":
            cell.alignment = Alignment(vertical="center", horizontal="right")

    ws.row_dimensions[excel_row].height = 18


# ============================================================
#  CO Document Index registration
# ============================================================

def _resolve_co_for_snapshot(snapshot_name):
    """Find the CO that this snapshot was used in. The CO has a
    `snapshot` field linking back to the snapshot."""
    return frappe.db.get_value(
        "InductOne Configuration Order",
        {"snapshot": snapshot_name},
        "name",
    )


def _register_in_co_document_index(co_name, snapshot_name, file_url):
    """
    Add or update the Document Index row for the hierarchy workbook on
    the linked CO. Mirrors the pattern in bom_export.sync_package_into_configuration_order.
    """
    co = frappe.get_doc("InductOne Configuration Order", co_name)

    row_title = "Configured BOM Hierarchy - {0}".format(snapshot_name)
    row_note = (
        "Indented BOM as resolved against the configured snapshot. "
        "Authoritative for the released configuration. "
        "Snapshot: {0}".format(snapshot_name)
    )

    existing_row = None
    for row in (co.documents or []):
        title = (getattr(row, "doc_title", "") or "").strip()
        if "Configured BOM Hierarchy" in title and snapshot_name in title:
            existing_row = row
            break

    if existing_row:
        existing_row.source_type = "MANUAL"
        existing_row.source_name = snapshot_name
        existing_row.doc_type = "OTHER"
        existing_row.doc_title = row_title
        existing_row.file_url = file_url
        if hasattr(existing_row, "file"):
            existing_row.file = file_url
        existing_row.required = "YES"
        existing_row.sort_order = 250  # Between CO PDF (~100) and BOM Export Package (~300)
        existing_row.small_text_vtsj = row_note
    else:
        co.append("documents", {
            "source_type": "MANUAL",
            "source_name": snapshot_name,
            "doc_type": "OTHER",
            "doc_title": row_title,
            "file": file_url,
            "file_url": file_url,
            "required": "YES",
            "sort_order": 250,
            "small_text_vtsj": row_note,
        })

    co.save(ignore_permissions=True)
