"""
================================================================================
 SNAPSHOT DIFF -- FRAPPE LOADER & PUBLIC API
================================================================================

The only module in the diff tool that imports frappe. It loads Configured BOM
Snapshot records, runs them through the schema-contract adapter (schema.py) to
produce normalized SnapshotNode lists, hands them to the pure diff engine
(engine.py), and exposes whitelisted methods for the UI.

Public whitelisted methods:
  get_diff(snapshot_a, snapshot_b, include_unchanged=0)
      -> structured dict for on-screen / programmatic use
  download_diff_workbook(snapshot_a, snapshot_b, include_unchanged=0)
      -> generates a color-coded XLSX and returns it as a download
         (does NOT attach to any record, by design)

If snapshot storage ever changes, you do NOT touch this file for field
renames -- that is schema.py's job. You only touch this file if the *way you
load* a snapshot changes (e.g. the parentfield name, or the snapshot doctype).
================================================================================
"""

import io

import frappe
from frappe import _
from frappe.utils import now_datetime

from .schema import (
    normalize_row,
    assert_schema_compatible,
    HIERARCHY_PARENTFIELD,
    SNAPSHOT_DOCTYPE,
    SNAPSHOT_SCHEMA_VERSION,
)
from .engine import (
    diff_snapshots,
    ADDED, REMOVED, QTY_CHANGED, REVISION_CHANGED, MOVED, UNCHANGED,
)
from .tree import (
    diff_snapshots_tree, flatten_tree, CHANGED,
)


# ----------------------------------------------------------------------------
#  Loading
# ----------------------------------------------------------------------------

def load_snapshot_nodes(snapshot_name):
    """
    Load one Configured BOM Snapshot's hierarchy as a list of normalized
    SnapshotNode objects. All field access is funneled through schema.py.
    """
    if not snapshot_name:
        frappe.throw(_("A snapshot name is required."))

    if not frappe.db.exists(SNAPSHOT_DOCTYPE, snapshot_name):
        frappe.throw(_("Configured BOM Snapshot '{0}' does not exist.").format(snapshot_name))

    snap = frappe.get_doc(SNAPSHOT_DOCTYPE, snapshot_name)

    # Fail loudly if the storage schema no longer matches the contract.
    assert_schema_compatible(snap)

    rows = snap.get(HIERARCHY_PARENTFIELD) or []
    nodes = [normalize_row(r.as_dict() if hasattr(r, "as_dict") else dict(r)) for r in rows]
    return nodes


def _snapshot_label(snapshot_name):
    """Human label for a snapshot: include the build and generation time if available."""
    meta = frappe.db.get_value(
        SNAPSHOT_DOCTYPE, snapshot_name,
        ["inductone_build", "generated_at", "top_bom"],
        as_dict=True,
    ) or {}
    parts = [snapshot_name]
    if meta.get("inductone_build"):
        parts.append("build {0}".format(meta["inductone_build"]))
    if meta.get("generated_at"):
        parts.append(str(meta["generated_at"]))
    return " | ".join(parts)


# ----------------------------------------------------------------------------
#  Public API -- structured diff
# ----------------------------------------------------------------------------

@frappe.whitelist()
def get_diff(snapshot_a, snapshot_b, include_unchanged=0):
    """
    Compute the diff between two snapshots and return a JSON-able dict.

    snapshot_a is treated as the OLDER / reference build.
    snapshot_b is treated as the NEWER / target build.
    """
    include_unchanged = int(include_unchanged or 0)

    nodes_a = load_snapshot_nodes(snapshot_a)
    nodes_b = load_snapshot_nodes(snapshot_b)

    result = diff_snapshots(
        nodes_a, nodes_b, snapshot_a, snapshot_b,
        include_unchanged=bool(include_unchanged),
    )

    return {
        "snapshot_a": result.snapshot_a,
        "snapshot_b": result.snapshot_b,
        "snapshot_a_label": _snapshot_label(snapshot_a),
        "snapshot_b_label": _snapshot_label(snapshot_b),
        "schema_version": result.schema_version,
        "summary": {
            "added": result.added,
            "removed": result.removed,
            "qty_changed": result.qty_changed,
            "revision_changed": result.revision_changed,
            "moved": result.moved,
            "unchanged": result.unchanged,
            "total_changes": result.total_changes,
        },
        "lines": [
            {
                "item_code": ln.item_code,
                "item_name": ln.item_name,
                "item_group": ln.item_group,
                "categories": ln.categories,
                "primary_category": ln.primary_category,
                "a_qty": ln.a_qty,
                "b_qty": ln.b_qty,
                "a_revision": ln.a_revision,
                "b_revision": ln.b_revision,
                "a_bom": ln.a_bom,
                "b_bom": ln.b_bom,
                "a_parent": ln.a_parent,
                "b_parent": ln.b_parent,
                "uom": ln.b_uom or ln.a_uom,
                "note": ln.note,
            }
            for ln in result.lines
        ],
    }


# ----------------------------------------------------------------------------
#  Public API -- XLSX download
# ----------------------------------------------------------------------------

# Color scheme for the diff workbook -- aligned with the Operations Manual
# brand palette and standard red/green/amber change semantics.
_FILL_ADDED = "DCFCE7"        # green  -- new in B
_FILL_REMOVED = "FEE2E2"      # red    -- gone in B
_FILL_QTY = "FEF9C3"          # yellow -- quantity changed
_FILL_REVISION = "E0F2FE"     # blue   -- revision changed
_FILL_MOVED = "F3E8FF"        # purple -- relocated
_FILL_UNCHANGED = "F9FAFB"    # grey   -- unchanged (only if included)
_FILL_HEADER = "1794CE"       # brand blue header


def _category_fill(primary_category):
    return {
        ADDED: _FILL_ADDED,
        REMOVED: _FILL_REMOVED,
        QTY_CHANGED: _FILL_QTY,
        REVISION_CHANGED: _FILL_REVISION,
        MOVED: _FILL_MOVED,
        UNCHANGED: _FILL_UNCHANGED,
    }.get(primary_category, "FFFFFF")


@frappe.whitelist()
def download_diff_workbook(snapshot_a, snapshot_b, include_unchanged=0):
    """
    Generate a color-coded diff XLSX and stream it as a download.

    Per design decision: this does NOT attach the workbook to either snapshot
    or any other record. It returns the file as a one-off download. The diff is
    a derived, on-demand artifact, not a record of truth -- the snapshots
    themselves are the records of truth.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    include_unchanged = int(include_unchanged or 0)

    nodes_a = load_snapshot_nodes(snapshot_a)
    nodes_b = load_snapshot_nodes(snapshot_b)
    result = diff_snapshots(
        nodes_a, nodes_b, snapshot_a, snapshot_b,
        include_unchanged=bool(include_unchanged),
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Snapshot Diff"

    thin = Border(bottom=Side(style="thin", color="E0E0E0"))

    # ---- Provenance / summary block ----
    ws.cell(row=1, column=1, value="Configured Snapshot Diff").font = Font(bold=True, size=16, color="1F2A44")

    meta = [
        ("Previous build (A)", _snapshot_label(snapshot_a)),
        ("This build (B)", _snapshot_label(snapshot_b)),
        ("Schema version", result.schema_version),
        ("Generated", str(now_datetime())),
        ("", ""),
        ("Added", result.added),
        ("Removed", result.removed),
        ("Revision changed", result.revision_changed),
        ("Qty changed", result.qty_changed),
        ("Moved", result.moved),
        ("Total changes", result.total_changes),
    ]
    for i, (label, value) in enumerate(meta, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True, color="555555", size=10)
        ws.cell(row=i, column=2, value=value).font = Font(color="333333", size=10)

    # ---- Legend ----
    legend_row = 3
    legend_col = 4
    legend = [
        ("ADDED — must be added this build", _FILL_ADDED),
        ("REMOVED — do NOT include this build", _FILL_REMOVED),
        ("REVISION CHANGED — build to new revision", _FILL_REVISION),
        ("QTY CHANGED — different quantity", _FILL_QTY),
        ("MOVED — different assembly", _FILL_MOVED),
    ]
    for i, (label, fill) in enumerate(legend):
        c = ws.cell(row=legend_row + i, column=legend_col, value=label)
        c.fill = PatternFill("solid", fgColor=fill)
        c.font = Font(size=10, color="333333")
        c.alignment = Alignment(horizontal="left", indent=1)
        ws.merge_cells(start_row=legend_row + i, start_column=legend_col,
                       end_row=legend_row + i, end_column=legend_col + 3)

    # ---- Column headers ----
    header_row = 16
    columns = [
        ("Change", "primary_category", 18),
        ("Item Code", "item_code", 20),
        ("Item Name", "item_name", 30),
        ("Item Group", "item_group", 18),
        ("Qty (A→B)", "_qty", 14),
        ("Revision (A→B)", "_rev", 18),
        ("Parent (A→B)", "_parent", 28),
        ("What to do", "note", 50),
    ]
    for col_idx, (label, _key, width) in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=col_idx, value=label)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor=_FILL_HEADER)
        c.alignment = Alignment(vertical="center")
        c.border = thin
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # ---- Data rows ----
    r = header_row + 1
    for ln in result.lines:
        fill = _category_fill(ln.primary_category)

        def _qty_str():
            if ln.a_qty is None:
                return "— → {0}".format(_fmt_qty(ln.b_qty))
            if ln.b_qty is None:
                return "{0} → —".format(_fmt_qty(ln.a_qty))
            if abs((ln.a_qty or 0) - (ln.b_qty or 0)) > 1e-9:
                return "{0} → {1}".format(_fmt_qty(ln.a_qty), _fmt_qty(ln.b_qty))
            return _fmt_qty(ln.b_qty)

        def _rev_str():
            a = ln.a_revision or "—"
            b = ln.b_revision or "—"
            if a != b:
                return "{0} → {1}".format(a, b)
            return b

        def _parent_str():
            a = ln.a_parent or "—"
            b = ln.b_parent or "—"
            if a != b:
                return "{0} → {1}".format(a, b)
            return b

        values = {
            "primary_category": ln.primary_category.replace("_", " ").title(),
            "item_code": ln.item_code,
            "item_name": ln.item_name,
            "item_group": ln.item_group,
            "_qty": _qty_str(),
            "_rev": _rev_str(),
            "_parent": _parent_str(),
            "note": ln.note,
        }
        for col_idx, (_label, key, _w) in enumerate(columns, start=1):
            c = ws.cell(row=r, column=col_idx, value=values.get(key, ""))
            c.fill = PatternFill("solid", fgColor=fill)
            c.border = thin
            c.alignment = Alignment(vertical="center", wrap_text=(key == "note"))
            c.font = Font(size=10, color="333333")
        r += 1

    if r > header_row + 1:
        ws.auto_filter.ref = "A{0}:{1}{2}".format(
            header_row, get_column_letter(len(columns)), r - 1
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = "Snapshot_Diff_{0}_vs_{1}_{2}.xlsx".format(
        snapshot_a, snapshot_b, now_datetime().strftime("%Y%m%d_%H%M%S")
    )

    frappe.local.response.filename = fname
    frappe.local.response.filecontent = buf.read()
    frappe.local.response.type = "download"


def _fmt_qty(q):
    if q is None:
        return "—"
    if abs(q - round(q)) < 1e-9:
        return str(int(round(q)))
    return "{0:.3f}".format(q).rstrip("0").rstrip(".")


# ============================================================================
#  REPORT DATA -- the method the UI Script Report calls
# ============================================================================
#
#  The pasted Script Report runs inside Frappe's safe_exec sandbox, which
#  forbids imports and any name starting with an underscore. So the report
#  script is a thin shell that calls THIS method via frappe.call(). All the
#  real work happens here, in unrestricted app code.
#
#  Returns a dict: {"columns": [...], "data": [...], "message": "...",
#                   "report_summary": [...]}
#  The shell hands columns/data/message/report_summary straight back to Frappe.

@frappe.whitelist()
def get_report_data(snapshot_a=None, snapshot_b=None,
                    view_mode="Hierarchical", context_mode="Changes only"):
    """
    Build the full report payload for the Configured Snapshot Diff report.

    Called from the UI Script Report shell. All logic lives here (unrestricted
    app code) rather than in the sandboxed report script.
    """
    if not snapshot_a or not snapshot_b:
        return {
            "columns": _report_empty_columns(view_mode),
            "data": [],
            "message": _report_intro_message(),
            "report_summary": [],
        }

    if snapshot_a == snapshot_b:
        return {
            "columns": _report_empty_columns(view_mode),
            "data": [],
            "message": _(
                "Both snapshots are the same -- there is nothing to diff."
            ),
            "report_summary": [],
        }

    nodes_a = load_snapshot_nodes(snapshot_a)
    nodes_b = load_snapshot_nodes(snapshot_b)

    if view_mode == "Flat Procurement":
        return _report_flat(nodes_a, nodes_b, snapshot_a, snapshot_b, context_mode)
    return _report_tree(nodes_a, nodes_b, snapshot_a, snapshot_b, context_mode)


def _report_tree(nodes_a, nodes_b, snapshot_a, snapshot_b, context_mode):
    changes_only = (context_mode == "Changes only")
    res = diff_snapshots_tree(
        nodes_a, nodes_b, snapshot_a, snapshot_b, changes_only=changes_only
    )

    columns = [
        {"label": _("Change"), "fieldname": "status", "fieldtype": "Data", "width": 110},
        {"label": _("Item (indented by level)"), "fieldname": "item_display", "fieldtype": "Data", "width": 360},
        {"label": _("Item Name"), "fieldname": "item_name", "fieldtype": "Data", "width": 220},
        {"label": _("Qty A to B"), "fieldname": "qty_change", "fieldtype": "Data", "width": 110},
        {"label": _("Revision A to B"), "fieldname": "rev_change", "fieldtype": "Data", "width": 140},
        {"label": _("Level"), "fieldname": "bom_level", "fieldtype": "Int", "width": 60},
        {"label": _("What changed"), "fieldname": "note", "fieldtype": "Data", "width": 360},
    ]

    indent_unit = "\u00a0\u00a0\u00a0\u00a0"
    marker_map = {ADDED: "[+] ", REMOVED: "[-] ", CHANGED: "[~] ", UNCHANGED: "    "}

    data = []
    for n in flatten_tree(res):
        indent = indent_unit * n.bom_level
        marker = marker_map.get(n.status, "")
        data.append({
            "status": n.status.title(),
            "item_display": "{0}{1}{2}".format(marker, indent, n.item_code),
            "item_name": n.item_name,
            "qty_change": _report_pair(n.a_qty, n.b_qty),
            "rev_change": _report_rev_pair(n.a_revision, n.b_revision),
            "bom_level": n.bom_level,
            "note": n.note,
            "_status_raw": n.status,
        })

    return {
        "columns": columns,
        "data": data,
        "message": _report_banner("Hierarchical", snapshot_a, snapshot_b, context_mode),
        "report_summary": [
            {"label": _("Added"), "value": res.added, "indicator": "Green"},
            {"label": _("Removed"), "value": res.removed, "indicator": "Red"},
            {"label": _("Changed"), "value": res.changed, "indicator": "Blue"},
            {"label": _("Unchanged"), "value": res.unchanged, "indicator": "Grey"},
            {"label": _("Total Changes"), "value": res.total_changes, "indicator": "Orange"},
        ],
    }


def _report_flat(nodes_a, nodes_b, snapshot_a, snapshot_b, context_mode):
    include_unchanged = (context_mode == "Show full list")
    res = diff_snapshots(
        nodes_a, nodes_b, snapshot_a, snapshot_b,
        include_unchanged=include_unchanged,
    )

    columns = [
        {"label": _("Change"), "fieldname": "category", "fieldtype": "Data", "width": 150},
        {"label": _("Item Code"), "fieldname": "item_code", "fieldtype": "Link", "options": "Item", "width": 160},
        {"label": _("Item Name"), "fieldname": "item_name", "fieldtype": "Data", "width": 240},
        {"label": _("Item Group"), "fieldname": "item_group", "fieldtype": "Data", "width": 150},
        {"label": _("Total Qty A to B"), "fieldname": "qty_change", "fieldtype": "Data", "width": 130},
        {"label": _("Revision A to B"), "fieldname": "rev_change", "fieldtype": "Data", "width": 150},
        {"label": _("UOM"), "fieldname": "uom", "fieldtype": "Data", "width": 70},
        {"label": _("What to do"), "fieldname": "note", "fieldtype": "Data", "width": 380},
    ]

    data = []
    for ln in res.lines:
        data.append({
            "category": ln.primary_category.replace("_", " ").title(),
            "item_code": ln.item_code,
            "item_name": ln.item_name,
            "item_group": ln.item_group,
            "qty_change": _report_pair(ln.a_qty, ln.b_qty),
            "rev_change": _report_rev_pair(ln.a_revision, ln.b_revision),
            "uom": ln.b_uom or ln.a_uom,
            "note": ln.note,
            "_status_raw": ln.primary_category,
        })

    return {
        "columns": columns,
        "data": data,
        "message": _report_banner("Flat Procurement", snapshot_a, snapshot_b, context_mode),
        "report_summary": [
            {"label": _("Added"), "value": res.added, "indicator": "Green"},
            {"label": _("Removed"), "value": res.removed, "indicator": "Red"},
            {"label": _("Revision Changed"), "value": res.revision_changed, "indicator": "Blue"},
            {"label": _("Qty Changed"), "value": res.qty_changed, "indicator": "Orange"},
            {"label": _("Moved"), "value": res.moved, "indicator": "Purple"},
            {"label": _("Total Changes"), "value": res.total_changes, "indicator": "Orange"},
        ],
    }


def _report_pair(a, b):
    if a is None and b is None:
        return ""
    if a is None:
        return "(none) -> {0}".format(_report_fmt(b))
    if b is None:
        return "{0} -> (none)".format(_report_fmt(a))
    if abs((a or 0) - (b or 0)) > 1e-9:
        return "{0} -> {1}".format(_report_fmt(a), _report_fmt(b))
    return _report_fmt(b)


def _report_rev_pair(a, b):
    a = a or ""
    b = b or ""
    if a != b:
        return "{0} -> {1}".format(a or "(none)", b or "(none)")
    return b


def _report_fmt(q):
    if q is None:
        return "(none)"
    try:
        q = float(q)
    except (TypeError, ValueError):
        return str(q)
    if abs(q - round(q)) < 1e-9:
        return str(int(round(q)))
    return "{0:.3f}".format(q).rstrip("0").rstrip(".")


def _report_banner(view, a, b, context_mode):
    # A swatch chip for the legend.
    def chip(color, label):
        return (
            "<span style='display:inline-block;padding:2px 10px;margin:2px 4px;"
            "border-radius:4px;background:{0};color:#1f2a44;font-size:0.85em;"
            "white-space:nowrap;'>{1}</span>"
        ).format(color, label)

    header = (
        "<div style='margin-bottom:8px;'>"
        "<b>{view} view</b> &mdash; comparing <b>{a}</b> (previous build) "
        "against <b>{b}</b> (this build). "
        "<span style='color:#667085;'>Context: {ctx}.</span>"
        "</div>"
    ).format(view=view, a=a, b=b, ctx=context_mode)

    purpose = (
        "<div style='margin-bottom:10px;color:#475467;'>"
        "This is a comparison of two configured builds. It exists so that "
        "knowledge from the previous build can be <b>amended, not blindly "
        "reused</b> &mdash; only the lines below are different between the two "
        "builds. Everything not shown is unchanged and can be built the same "
        "way as before."
        "</div>"
    )

    # Legend chips -- same colors as the rows.
    legend = (
        "<div style='margin-bottom:8px;'>"
        + chip("#DCFCE7", "Added")
        + chip("#FEE2E2", "Removed")
        + chip("#E0F2FE", "Revision changed")
        + chip("#FEF9C3", "Qty changed")
        + chip("#F3E8FF", "Moved")
        + "</div>"
    )

    # Glossary -- the non-obvious meanings, tailored to the active view.
    if view == "Flat Procurement":
        glossary_rows = [
            ("Added",
             "A part that was NOT on the previous build and IS on this one. "
             "You must source/buy it for this build."),
            ("Removed",
             "A part that WAS on the previous build and is NOT on this one. "
             "Do not buy or include it this time."),
            ("Revision changed",
             "The same part, but it is now sourced from a different BOM "
             "revision than before. The part number may be identical &mdash; "
             "what changed is the revision of the bill of materials it comes "
             "from. Treat it as: re-verify this part against the new revision."),
            ("Qty changed",
             "The same part, but the <b>total</b> quantity needed across the "
             "whole build is different. The Total Qty column shows "
             "<i>previous &rarr; this build</i>."),
            ("Revision A &rarr; B column",
             "Shows the BOM revision(s) the part is sourced from, previous "
             "build on the left, this build on the right. If a part is used "
             "in more than one BOM, you may see more than one revision listed "
             "&mdash; that means the part lives in multiple bills of material, "
             "each at its own revision. That is normal for shared/common parts."),
        ]
    else:
        glossary_rows = [
            ("&#10133; Added",
             "This node is new in this build &mdash; it did not exist at this "
             "position in the previous build's tree."),
            ("&#10134; Removed",
             "This node was in the previous build at this position and is gone "
             "in this one."),
            ("&#9999; Changed",
             "This node exists in both builds at the same position, but its "
             "quantity and/or its BOM revision changed. The What-changed "
             "column spells out which."),
            ("Indentation / Level",
             "Rows are indented by their depth in the assembly. A change shown "
             "under a parent assembly means that change happens <i>inside</i> "
             "that assembly. Parent rows are kept (even when unchanged) so you "
             "can see where in the structure each change lives."),
            ("Revision changed",
             "The part is sourced from a different BOM revision than before. "
             "Per the rule we use: a row is flagged only when <b>its own</b> "
             "sourcing BOM revision changes. A parent assembly's revision "
             "bumping does NOT flag its children unless the child itself "
             "actually changed &mdash; so if a child is unflagged, its use is "
             "genuinely unchanged and prior knowledge still applies."),
        ]

    glossary = (
        "<details style='margin-top:4px;'>"
        "<summary style='cursor:pointer;color:#1794ce;font-weight:600;'>"
        "What do these mean? (click to expand)</summary>"
        "<table style='margin-top:8px;border-collapse:collapse;width:100%;'>"
    )
    for i, (term, desc) in enumerate(glossary_rows):
        bg = "#f9fafb" if i % 2 == 0 else "#ffffff"
        glossary += (
            "<tr style='background:{0};'>"
            "<td style='padding:6px 10px;vertical-align:top;font-weight:600;"
            "white-space:nowrap;border-bottom:1px solid #eaecf0;'>{1}</td>"
            "<td style='padding:6px 10px;vertical-align:top;color:#475467;"
            "border-bottom:1px solid #eaecf0;'>{2}</td>"
            "</tr>"
        ).format(bg, term, desc)
    glossary += "</table></details>"

    note = (
        "<div style='margin-top:8px;color:#667085;font-size:0.85em;'>"
        "Switch <b>View Mode</b> for the engineering tree vs. the flat "
        "procurement list. Use <b>Download Diff Workbook</b> for a color-coded "
        "export you can send to the builder."
        "</div>"
    )

    return header + purpose + legend + glossary + note


def _report_intro_message():
    return _(
        "Select <b>Snapshot A</b> (the previous/reference build) and "
        "<b>Snapshot B</b> (this build) to compare. The diff shows exactly "
        "what changed between the two configurations so prior build knowledge "
        "can be amended rather than blindly reused."
    )


def _report_empty_columns(view_mode):
    if view_mode == "Flat Procurement":
        return [
            {"label": _("Change"), "fieldname": "category", "fieldtype": "Data", "width": 150},
            {"label": _("Item Code"), "fieldname": "item_code", "fieldtype": "Data", "width": 160},
        ]
    return [
        {"label": _("Change"), "fieldname": "status", "fieldtype": "Data", "width": 110},
        {"label": _("Item"), "fieldname": "item_display", "fieldtype": "Data", "width": 360},
    ]