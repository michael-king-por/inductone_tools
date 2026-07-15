#!/usr/bin/env python3
"""Backfill POR locations, legacy Instances, and FCO records from seed files.

This script is candidate/deployment maintenance tooling. It ingests the seed
workbooks from ``scripts/inductone_backfill/seeds`` without re-deriving their
data and writes a JSON evidence summary.
"""

from __future__ import annotations

import argparse
import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import frappe
from openpyxl import load_workbook


REQUIRED_CUSTOMERS = ["UPS", "DHL", "Amazon", "Plus One Robotics"]
LOCATION_DOCTYPE = "POR Physical Location"
INSTANCE_DOCTYPE = "InductOne Instance"
REQUEST_DOCTYPE = "InductOne Field Change Request"
FIELD_CHANGE_DOCTYPE = "InductOne Field Change"


def rows(path: Path, sheet: str | None = None) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    iterator = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    out = []
    for row in iterator:
        if not any(value is not None for value in row):
            continue
        out.append(dict(zip(headers, row)))
    return out


def parse_date(value):
    if not value:
        return None
    if hasattr(value, "date"):
        return value.date()
    for fmt in ("%b %d, %Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(str(value), fmt).date()
        except ValueError:
            pass
    return None


def phase0_customers() -> dict[str, bool]:
    return {name: bool(frappe.db.exists("Customer", name)) for name in REQUIRED_CUSTOMERS}


def naming_series(location_type: str) -> str:
    return {
        "Customer": "POR-CUST.####",
        "Site": "POR-SITE.####",
        "Lane": "POR-LANE.####",
        "Cell": "POR-CELL.####",
        "Robot": "POR-ROBOT.####",
    }[location_type]


def location_name_by_code(code: str | None) -> str | None:
    if not code:
        return None
    return frappe.db.get_value(LOCATION_DOCTYPE, {"location_code": code}, "name")


def seed_locations(seed_dir: Path) -> dict[str, Any]:
    seeded = rows(seed_dir / "location_tree_seed.xlsx", "POR Physical Location")
    created, existing = [], []
    ensure_customer_roots(seeded, created, existing)
    pending = seeded[:]

    # Parent-before-child loop. If a parent is missing after a full pass, that is
    # a seed conflict and should fail loudly.
    while pending:
        progressed = False
        next_pending = []
        for row in pending:
            code = row["location_code"]
            if location_name_by_code(code):
                existing.append(code)
                restamp_location_from_seed(row)
                progressed = True
                continue

            parent_code = row.get("parent_location_code")
            if row.get("location_type") == "Site" and not parent_code:
                parent_code = row.get("customer")
            parent_name = location_name_by_code(parent_code) if parent_code else None
            if parent_code and not parent_name:
                next_pending.append(row)
                continue

            doc = frappe.get_doc(
                {
                    "doctype": LOCATION_DOCTYPE,
                    "naming_series": naming_series(row["location_type"]),
                    "location_type": row["location_type"],
                    "location_code": code,
                    "location_name": row["location_name"],
                    "customer": row.get("customer"),
                    "site_code": row.get("site_code"),
                    "lane_code": row.get("lane_code"),
                    "cell_code": row.get("cell_code"),
                    "full_path": row.get("full_path"),
                    "is_group": 1 if row.get("is_group") else 0,
                    "parent_por_physical_location": parent_name,
                    "old_parent": parent_name,
                }
            )
            doc.insert(ignore_permissions=True)
            restamp_location_from_seed(row)
            created.append(code)
            progressed = True

        if next_pending and not progressed:
            raise RuntimeError(f"Unresolved location parents: {next_pending}")
        pending = next_pending

    return {"created": created, "existing": existing, "seed_rows": len(seeded)}


def ensure_customer_roots(
    seeded: list[dict[str, Any]],
    created: list[str],
    existing: list[str],
) -> None:
    """Create one Customer group node per distinct seed customer."""

    for customer in sorted({row.get("customer") for row in seeded if row.get("customer")}):
        if location_name_by_code(customer):
            existing.append(customer)
            continue

        doc = frappe.get_doc(
            {
                "doctype": LOCATION_DOCTYPE,
                "naming_series": naming_series("Customer"),
                "location_type": "Customer",
                "location_code": customer,
                "location_name": customer,
                "customer": customer,
                "full_path": customer,
                "is_group": 1,
            }
        )
        doc.insert(ignore_permissions=True)
        created.append(customer)


def restamp_location_from_seed(row: dict[str, Any]) -> None:
    """Preserve seed-facing location labels after Frappe tree insertion.

    Frappe's tree/nested-set helpers may derive labels from document names/codes
    during insert. The seed's full_path is the CSA-facing human label we need on
    the Instance and reports, so backfill re-stamps it idempotently.
    """

    name = location_name_by_code(row.get("location_code"))
    if not name:
        return

    parent_code = row.get("parent_location_code")
    if row.get("location_type") == "Site" and not parent_code:
        parent_code = row.get("customer")
    parent_name = location_name_by_code(parent_code) if parent_code else None

    doc = frappe.get_doc(LOCATION_DOCTYPE, name)
    if parent_name and doc.parent_por_physical_location != parent_name:
        doc.parent_por_physical_location = parent_name
        doc.save(ignore_permissions=True)

    full_path = row.get("full_path")
    if row.get("customer") and full_path and not str(full_path).startswith(str(row.get("customer")) + " /"):
        full_path = f"{row.get('customer')} / {full_path}"

    frappe.db.set_value(
        LOCATION_DOCTYPE,
        name,
        {
            "full_path": full_path,
            "is_group": 1 if row.get("is_group") else 0,
        },
        update_modified=False,
    )


def normalize_origin(value: str | None) -> str:
    value = value or ""
    if value.startswith("POC"):
        return "POC"
    if "Internal" in value:
        return "Internal-Reference"
    if value == "Legacy backfill":
        return "Legacy backfill"
    return value or "Legacy backfill"


def seed_instances(seed_dir: Path) -> dict[str, Any]:
    instance_rows = rows(seed_dir / "instance_backfill_seed.xlsx", "Instances")
    serial_rows = rows(seed_dir / "instance_backfill_seed.xlsx", "Component Serials")
    pending_rows = rows(seed_dir / "instance_backfill_seed.xlsx", "Pending Serial")
    serials_by_instance = defaultdict(list)
    for serial in serial_rows:
        serials_by_instance[serial["instance_system_serial"]].append(serial)

    created, existing, skipped_pending = [], [], [r["provisional_id"] for r in pending_rows]
    for row in instance_rows:
        system_serial = row["system_serial"]
        existing_name = frappe.db.exists(INSTANCE_DOCTYPE, system_serial)
        if existing_name:
            update_existing_instance_from_seed(existing_name, row, serials_by_instance.get(system_serial, []))
            existing.append(system_serial)
            continue

        physical_location = location_name_by_code(row.get("physical_location"))
        if row.get("physical_location") and not physical_location:
            raise RuntimeError(f"Instance {system_serial} references missing location {row.get('physical_location')}")

        doc = frappe.get_doc(
            {
                "doctype": INSTANCE_DOCTYPE,
                "system_serial": system_serial,
                "origin": normalize_origin(row.get("origin")),
                "status": row.get("status"),
                "customer": row.get("customer"),
                "deployment_site": row.get("deployment_site"),
                "customer_project_label": row.get("customer_project_label"),
                "orientation": row.get("orientation"),
                "build_date": parse_date(row.get("build_date")),
                "physical_location": physical_location,
                "configuration_summary": row.get("configuration_summary"),
                "backfill_notes": row.get("backfill_notes"),
            }
        )
        for serial in serials_by_instance.get(system_serial, []):
            doc.append(
                "component_serials",
                {
                    "component_label": serial.get("component_type"),
                    "serial_number": serial.get("serial"),
                },
            )
        doc.insert(ignore_permissions=True)
        created.append(system_serial)

    return {
        "created": created,
        "existing": existing,
        "component_serial_rows": len(serial_rows),
        "pending_serials_skipped": skipped_pending,
        "seed_rows": len(instance_rows),
    }


def seed_instance_values(row: dict[str, Any]) -> dict[str, Any]:
    physical_location = location_name_by_code(row.get("physical_location"))
    if row.get("physical_location") and not physical_location:
        raise RuntimeError(f"Instance {row['system_serial']} references missing location {row.get('physical_location')}")
    deployment_site = frappe.db.get_value(LOCATION_DOCTYPE, physical_location, "full_path") if physical_location else row.get("deployment_site")
    return {
        "origin": normalize_origin(row.get("origin")),
        "status": row.get("status"),
        "customer": row.get("customer"),
        "deployment_site": deployment_site,
        "customer_project_label": row.get("customer_project_label"),
        "orientation": row.get("orientation"),
        "build_date": parse_date(row.get("build_date")),
        "physical_location": physical_location,
        "configuration_summary": row.get("configuration_summary"),
        "backfill_notes": row.get("backfill_notes"),
    }


def update_existing_instance_from_seed(instance_name: str, row: dict[str, Any], serial_rows: list[dict[str, Any]]) -> None:
    # This is maintenance/backfill tooling. Existing candidate records may have
    # been partially created before a previous failed run, including with a
    # lifecycle status that the normal Desk validator would not allow to move
    # backwards. Stamp seed scalar fields directly, then reload for any child
    # table additions.
    frappe.db.set_value(
        INSTANCE_DOCTYPE,
        instance_name,
        seed_instance_values(row),
        update_modified=False,
    )
    doc = frappe.get_doc(INSTANCE_DOCTYPE, instance_name)
    if not doc.get("component_serials") and serial_rows:
        for serial in serial_rows:
            doc.append(
                "component_serials",
                {
                    "component_label": serial.get("component_type"),
                    "serial_number": serial.get("serial"),
                },
            )
    doc.save(ignore_permissions=True)


def disposition_from_flow(flow_status: str | None) -> tuple[str, str]:
    value = (flow_status or "").strip().lower()
    if value == "complete":
        return "Closed", "Approved"
    if value == "denied":
        return "Rejected", "Rejected"
    if value in {"canceled", "cancelled"}:
        return "Cancelled", "Cancelled"
    return "Triaged", "Pending"


def triage_outcome(value: str | None) -> str:
    value = (value or "").strip()
    if value == "FCO":
        return "Field Change"
    if value in {"ECR", "Deviation"}:
        return value
    return value or "Field Change"


def fco_number_from_change_no(change_no: Any, map_rows: dict[str, dict[str, Any]]) -> str | None:
    if change_no is None:
        return None
    suffix = f"{int(change_no):03d}" if isinstance(change_no, (int, float)) else str(change_no).zfill(3)
    for fco_no in map_rows:
        if fco_no.endswith(suffix):
            return fco_no
    return None


def seed_requests_and_field_changes(seed_dir: Path) -> dict[str, Any]:
    map_list = rows(seed_dir / "fco_instance_map.xlsx", "FCO -> Instance Map")
    map_by_fco = {row["FCO #"]: row for row in map_list}
    jotform_rows = rows(seed_dir / "fco_jotform_export.xlsx", "Sheet1")

    created_requests, existing_requests, pending_or_organic = [], [], []
    request_names = []
    for jot in jotform_rows:
        fco_no = fco_number_from_change_no(jot.get("Change No"), map_by_fco)
        if not fco_no:
            raise RuntimeError(f"Could not map JotForm Change No {jot.get('Change No')} to fco_instance_map")
        mapped = map_by_fco[fco_no]
        request_names.append(fco_no)
        if frappe.db.exists(REQUEST_DOCTYPE, fco_no):
            existing_requests.append(fco_no)
            continue

        primary_serial = mapped.get("PRIMARY Unit (best guess) - serial")
        instance = primary_serial if primary_serial and frappe.db.exists(INSTANCE_DOCTYPE, primary_serial) else None
        if primary_serial and not instance:
            pending_or_organic.append({"fco": fco_no, "primary_serial": primary_serial, "reason": mapped.get("Basis for guess")})

        status, disposition = disposition_from_flow(jot.get("Flow Status"))
        doc = frappe.get_doc(
            {
                "doctype": REQUEST_DOCTYPE,
                "naming_series": "FCO-.YYYY.-.###",
                "intake_ref": fco_no,
                "status": status,
                "date_raised": parse_date(jot.get("Date") or jot.get("Submission Date")),
                "requester": jot.get("Change Requestor"),
                "requester_department": jot.get("Department"),
                "requester_role": jot.get("Role"),
                "intake_source": "JotForm Import",
                "title": jot.get("Title of Field Change Order") or mapped.get("Title"),
                "description": jot.get("Description of proposed change in plain English:"),
                "reason": jot.get("Benefit/Reason for change:"),
                "customer_project": jot.get("What site(s), project(s), or customer does this apply to?"),
                "machine_identifier": mapped.get("Primary Unit (label)") or primary_serial,
                "scope": mapped.get("Applicable Units"),
                "one_time_or_repeated": jot.get("Do you expect this to be a one-time fix, or repeated across sites?"),
                "est_downtime_h": jot.get("Estimated Customer Downtime hours"),
                "est_labor_h": jot.get("Estimated Labor Hours"),
                "parts_cost": jot.get("Part(s) Cost ($)"),
                "implementer": jot.get("Who can implement this field change?"),
                "tools_docs": jot.get("Tools / Documents Needed"),
                "ticket_link": jot.get("Relevant issue tracker link or ticket:"),
                "instance": instance,
                "triage_outcome": triage_outcome(mapped.get("Triage")),
                "reference": jot.get("ECR Link") or jot.get("Epic link"),
                "safety_regulatory": 0,
                "disposition": disposition,
                "disposition_date": parse_date(jot.get("Submittal Date")),
                "disposition_by": jot.get("Approver Email"),
                "assignment_confidence": normalize_confidence(mapped.get("Assignment Confidence")),
                "assignment_basis": mapped.get("Basis for guess"),
                "assignment_reviewed": 0,
                "notes": jot.get("Notes"),
            }
        )
        doc.insert(ignore_permissions=True)
        created_requests.append(fco_no)

    spawned, existing_fc = spawn_field_changes(map_by_fco)
    return {
        "requests_created": created_requests,
        "requests_existing": existing_requests,
        "field_changes_created": spawned,
        "field_changes_existing": existing_fc,
        "pending_or_organic": pending_or_organic,
        "map_rows": len(map_list),
        "jotform_rows": len(jotform_rows),
        "represented_requests": sorted(set(request_names)),
    }


def normalize_confidence(value: str | None) -> str:
    value = value or ""
    if value.startswith("High"):
        return "High"
    if value.startswith("Med"):
        return "Med"
    if value.startswith("Low"):
        return "Low"
    if "organic" in value.lower():
        return "Backfill-guess"
    return value or "Backfill-guess"


def spawn_field_changes(map_by_fco: dict[str, dict[str, Any]]) -> tuple[list[str], list[str]]:
    spawn_plan = {
        "FCO-2025-007": ["IN002024257003", "IN002024257004"],
        "FCO-2025-010": ["IN001024355007", "IN001024355008"],
    }
    created, existing = [], []
    for fco_no, serials in spawn_plan.items():
        mapped = map_by_fco[fco_no]
        for serial in serials:
            marker = f"{fco_no}-{serial}"
            found = frappe.db.exists(FIELD_CHANGE_DOCTYPE, {"source_request": fco_no, "instance": serial})
            if found:
                existing.append(found)
                continue
            if not frappe.db.exists(INSTANCE_DOCTYPE, serial):
                raise RuntimeError(f"Cannot spawn Field Change {marker}; missing Instance {serial}")

            doc = frappe.get_doc(
                {
                    "doctype": FIELD_CHANGE_DOCTYPE,
                    "naming_series": "FC-.YYYY.-.####",
                    "status": "Locked",
                    "instance": serial,
                    "source_request": fco_no,
                    "reference": None,
                    "change_summary": mapped.get("Title"),
                    "implemented_date": frappe.utils.today(),
                    "implemented_by": "Backfill",
                    "post_change_test": "N/A",
                    "as_maintained_updated": 1,
                    "accepted_by": frappe.session.user,
                    "accepted_at": frappe.utils.now_datetime(),
                    "notes": mapped.get("Basis for guess"),
                }
            )
            doc.insert(ignore_permissions=True)
            update_instance_rollup(serial)
            created.append(doc.name)
    return created, existing


def update_instance_rollup(serial: str) -> None:
    result = frappe.db.sql(
        """
        SELECT COUNT(*) AS count, MAX(implemented_date) AS latest_date
        FROM `tabInductOne Field Change`
        WHERE instance = %s AND status = 'Locked'
        """,
        serial,
        as_dict=True,
    )[0]
    frappe.db.set_value(
        INSTANCE_DOCTYPE,
        serial,
        {
            "field_change_count": result.count or 0,
            "latest_field_change_date": result.latest_date,
        },
        update_modified=False,
    )


def run(args) -> dict[str, Any]:
    frappe.init(site=args.site, sites_path=args.sites_path)
    frappe.connect()
    try:
        phase0 = phase0_customers()
        missing = [name for name, exists in phase0.items() if not exists]
        if missing:
            raise RuntimeError(f"Missing required Customers: {missing}")

        payload = {
            "site": args.site,
            "generated_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "phase0_customers": phase0,
            "locations": seed_locations(args.seed_dir),
            "instances": seed_instances(args.seed_dir),
            "fco": seed_requests_and_field_changes(args.seed_dir),
        }
        frappe.db.commit()
        return payload
    finally:
        frappe.destroy()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--site", required=True)
    parser.add_argument("--sites-path", required=True)
    parser.add_argument(
        "--seed-dir",
        type=Path,
        default=Path(__file__).resolve().parent / "seeds",
    )
    parser.add_argument(
        "--evidence-dir",
        type=Path,
        default=Path("/mnt/c/hub/frappe-sandbox/validation-evidence"),
    )
    args = parser.parse_args()
    args.evidence_dir.mkdir(parents=True, exist_ok=True)

    try:
        payload = run(args)
    except Exception as exc:
        payload = {"ok": False, "error": repr(exc)}
        path = args.evidence_dir / "fco_as_installed_backfill_failed.json"
        path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        print(json.dumps(payload, indent=2, default=str))
        return 1

    path = args.evidence_dir / "fco_as_installed_backfill.json"
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    print(json.dumps(payload, indent=2, default=str))
    print(f"Evidence: {path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
