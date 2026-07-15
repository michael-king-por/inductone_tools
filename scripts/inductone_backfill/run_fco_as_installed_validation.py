#!/usr/bin/env python3
"""Validate the InductOne as-installed/FCO integration gates in candidate."""

from __future__ import annotations

import argparse
import shutil
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import frappe
from frappe.permissions import has_permission
from openpyxl import load_workbook


LOCATION_DOCTYPE = "POR Physical Location"
INSTANCE_DOCTYPE = "InductOne Instance"
REQUEST_DOCTYPE = "InductOne Field Change Request"
FIELD_CHANGE_DOCTYPE = "InductOne Field Change"
NEW_DOCTYPES = [REQUEST_DOCTYPE, FIELD_CHANGE_DOCTYPE, "InductOne Field Change Serial"]


class GateRunner:
    def __init__(self):
        self.gates = []

    def check(self, gate: str, condition: bool, details=None):
        status = "PASS" if condition else "FAIL"
        self.gates.append({"gate": gate, "status": status, "details": details})
        print(status, gate, json.dumps(details, default=str) if details is not None else "")
        return condition

    @property
    def ok(self):
        return all(gate["status"] == "PASS" for gate in self.gates)


def workbook_rows(path: Path, sheet: str):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet]
    iterator = ws.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    return [dict(zip(headers, row)) for row in iterator if any(value is not None for value in row)]


def gate_location_tree(runner: GateRunner, seed_dir: Path):
    seed = workbook_rows(seed_dir / "location_tree_seed.xlsx", "POR Physical Location")
    missing, parent_errors, leaf_errors, path_errors, customer_errors = [], [], [], [], []

    for customer in sorted({row.get("customer") for row in seed if row.get("customer")}):
        customer_name = frappe.db.get_value(
            LOCATION_DOCTYPE,
            {"location_type": "Customer", "customer": customer},
            "name",
        )
        if not customer_name:
            customer_errors.append({"customer": customer, "issue": "missing customer root"})
            continue
        customer_doc = frappe.get_doc(LOCATION_DOCTYPE, customer_name)
        if customer_doc.parent_por_physical_location:
            customer_errors.append({"customer": customer, "issue": "customer root has parent", "parent": customer_doc.parent_por_physical_location})
        if customer_doc.full_path != customer:
            customer_errors.append({"customer": customer, "issue": "customer root full_path mismatch", "actual": customer_doc.full_path})

    for row in seed:
        code = row["location_code"]
        name = frappe.db.get_value(LOCATION_DOCTYPE, {"location_code": code}, "name")
        if not name:
            missing.append(code)
            continue
        doc = frappe.get_doc(LOCATION_DOCTYPE, name)
        parent_code = row.get("parent_location_code")
        if row.get("location_type") == "Site" and not parent_code:
            parent_code = row.get("customer")
        if parent_code:
            parent_name = frappe.db.get_value(LOCATION_DOCTYPE, {"location_code": parent_code}, "name")
            if doc.parent_por_physical_location != parent_name:
                parent_errors.append({"location_code": code, "expected_parent": parent_code, "actual": doc.parent_por_physical_location})
        if row["location_type"] == "Cell" and (int(doc.rgt or 0) - int(doc.lft or 0) != 1):
            leaf_errors.append({"location_code": code, "lft": doc.lft, "rgt": doc.rgt})
        expected_path = row.get("full_path")
        if row.get("customer") and expected_path and not str(expected_path).startswith(str(row.get("customer")) + " /"):
            expected_path = f"{row.get('customer')} / {expected_path}"
        if doc.full_path != expected_path:
            path_errors.append({"location_code": code, "expected": expected_path, "actual": doc.full_path})
    return runner.check(
        "Location tree integrity: Customer -> Site -> Lane -> Cell",
        not (missing or parent_errors or leaf_errors or path_errors or customer_errors),
        {
            "seed_rows": len(seed),
            "customer_roots": sorted({row.get("customer") for row in seed if row.get("customer")}),
            "missing": missing,
            "parent_errors": parent_errors,
            "leaf_errors": leaf_errors,
            "path_errors": path_errors,
            "customer_errors": customer_errors,
        },
    )


def gate_instance_locations(runner: GateRunner, seed_dir: Path):
    seed = workbook_rows(seed_dir / "instance_backfill_seed.xlsx", "Instances")
    seeded_serials = [row["system_serial"] for row in seed]
    rows = frappe.get_all(
        INSTANCE_DOCTYPE,
        filters={"name": ["in", seeded_serials]},
        fields=["name", "origin", "physical_location", "deployment_site"],
    )
    failures = []
    for row in rows:
        if row.origin == "Internal-Reference":
            continue
        if not row.physical_location:
            failures.append({"instance": row.name, "issue": "missing physical_location"})
            continue
        loc = frappe.get_doc(LOCATION_DOCTYPE, row.physical_location)
        if loc.location_type != "Cell":
            failures.append({"instance": row.name, "issue": "physical_location is not Cell", "location": row.physical_location, "type": loc.location_type})
        if row.deployment_site != loc.full_path:
            failures.append({"instance": row.name, "issue": "deployment_site mismatch", "expected": loc.full_path, "actual": row.deployment_site})
    missing = sorted(set(seeded_serials) - {row.name for row in rows})
    failures.extend({"instance": name, "issue": "seeded instance missing"} for name in missing)
    return runner.check("Seeded Instances linked to Cell physical locations", not failures, {"checked": len(rows), "seed_rows": len(seed), "failures": failures})


def gate_field_changes_resolve(runner: GateRunner):
    rows = frappe.get_all(FIELD_CHANGE_DOCTYPE, fields=["name", "instance"])
    failures = []
    for row in rows:
        instance = frappe.get_doc(INSTANCE_DOCTYPE, row.instance)
        if not instance.physical_location:
            failures.append({"field_change": row.name, "instance": row.instance, "issue": "instance missing physical_location"})
            continue
        loc = frappe.get_doc(LOCATION_DOCTYPE, instance.physical_location)
        if loc.location_type != "Cell":
            failures.append({"field_change": row.name, "instance": row.instance, "location": loc.name, "type": loc.location_type})
    return runner.check("Backfilled Field Changes resolve through Instance to Cell", not failures, {"checked": len(rows), "failures": failures})


def gate_fco_map_represented(runner: GateRunner, seed_dir: Path):
    seed = workbook_rows(seed_dir / "fco_instance_map.xlsx", "FCO -> Instance Map")
    missing_requests, missing_instances, spawned = [], [], []
    pending_or_organic = []
    for row in seed:
        fco = row["FCO #"]
        if not frappe.db.exists(REQUEST_DOCTYPE, fco):
            missing_requests.append(fco)
        serial = row.get("PRIMARY Unit (best guess) - serial")
        if serial and isinstance(serial, str) and serial.startswith(("IN", "IND", "REF")):
            if not frappe.db.exists(INSTANCE_DOCTYPE, serial):
                missing_instances.append({"fco": fco, "serial": serial})
        else:
            pending_or_organic.append({"fco": fco, "serial": serial})
        if str(row.get("Spawns per-Instance FC") or "").upper().startswith("YES"):
            spawned.append(fco)
    fc_counts = {
        fco: frappe.db.count(FIELD_CHANGE_DOCTYPE, {"source_request": fco})
        for fco in spawned
    }
    spawn_ok = fc_counts.get("FCO-2025-007") == 2 and fc_counts.get("FCO-2025-010") == 2
    return runner.check(
        "FCO map represented with pending/organic exceptions",
        not missing_requests and not missing_instances and spawn_ok,
        {
            "map_rows": len(seed),
            "missing_requests": missing_requests,
            "missing_instances": missing_instances,
            "spawn_counts": fc_counts,
            "pending_or_organic": pending_or_organic,
        },
    )


def gate_as_installed_query(runner: GateRunner):
    rows = frappe.db.sql(
        """
        SELECT
          site.location_code AS site_code,
          inst.name AS instance,
          inst.status,
          inst.latest_field_change_date
        FROM `tabInductOne Instance` inst
        INNER JOIN `tabPOR Physical Location` cell ON cell.name = inst.physical_location
        LEFT JOIN `tabPOR Physical Location` lane ON lane.name = cell.parent_por_physical_location
        LEFT JOIN `tabPOR Physical Location` site ON site.name = lane.parent_por_physical_location
        WHERE site.location_type = 'Site'
        ORDER BY site.location_code, inst.name
        """,
        as_dict=True,
    )
    return runner.check("As-installed Site query returns installed fleet context", bool(rows), {"row_count": len(rows), "sample": rows[:5]})


def gate_reassignment_mechanism(runner: GateRunner):
    request_name = frappe.db.get_value(REQUEST_DOCTYPE, {"assignment_confidence": "Low"}, "name")
    instances = frappe.get_all(INSTANCE_DOCTYPE, pluck="name", limit=2)
    if not request_name or len(instances) < 2:
        return runner.check("Reassignment mechanism", False, {"issue": "not enough test data"})

    doc = frappe.get_doc(REQUEST_DOCTYPE, request_name)
    original = doc.instance
    target = next(name for name in instances if name != original)
    before_versions = frappe.db.count("Version", {"ref_doctype": REQUEST_DOCTYPE, "docname": request_name})

    rejected_without_reason = False
    try:
        doc.instance = target
        doc.assignment_change_reason = ""
        doc.save()
    except Exception:
        rejected_without_reason = True
        frappe.db.rollback()

    doc = frappe.get_doc(REQUEST_DOCTYPE, request_name)
    doc.instance = target
    doc.assignment_change_reason = "Validation proof: reassignment requires reason."
    doc.save()
    doc = frappe.get_doc(REQUEST_DOCTYPE, request_name)
    doc.instance = original
    doc.assignment_change_reason = "Validation cleanup: restore original assignment."
    doc.save()
    frappe.db.commit()
    after_versions = frappe.db.count("Version", {"ref_doctype": REQUEST_DOCTYPE, "docname": request_name})

    low_conf_count = frappe.db.count(REQUEST_DOCTYPE, {"assignment_confidence": ["in", ["Low", "Backfill-guess"]], "assignment_reviewed": 0})
    return runner.check(
        "Reassignment mechanism rejects missing reason and records Version",
        rejected_without_reason and after_versions > before_versions and low_conf_count > 0,
        {"request": request_name, "before_versions": before_versions, "after_versions": after_versions, "low_confidence_unreviewed": low_conf_count},
    )


def gate_external_builder_denial(runner: GateRunner):
    users = ["motion.builder@plusonerobotics.com", "lam@plusonerobotics.com"]
    failures = []
    original_user = frappe.session.user
    for user in users:
        if not frappe.db.exists("User", user):
            failures.append({"user": user, "issue": "missing user"})
            continue
        frappe.set_user(user)
        for doctype in NEW_DOCTYPES:
            for ptype in ["read", "create"]:
                allowed = has_permission(doctype, ptype=ptype, user=user)
                if allowed:
                    failures.append({"user": user, "doctype": doctype, "ptype": ptype})
    frappe.set_user(original_user)
    return runner.check("External builders denied new FCO DocTypes", not failures, {"failures": failures})


def gate_ignore_permissions_clean(runner: GateRunner, app_path: Path):
    """Assert no whitelisted method bypasses permissions.

    Backfill/patch helpers are allowed to use ``ignore_permissions`` because they
    are console/migration tooling. The hardening rule we care about here is the
    user-callable boundary: a whitelisted method must gate permissions normally
    instead of silently bypassing them.
    """

    new_files = [
        app_path / "inductone_tools" / "field_change.py",
        app_path / "inductone_tools" / "physical_location.py",
        app_path / "inductone_tools" / "instance" / "backfill.py",
    ]
    hits = []
    for path in new_files:
        if not path.exists():
            hits.append({"file": str(path), "issue": "missing"})
            continue
        text = path.read_text(encoding="utf-8")
        whitelist_lines = [match.start() for match in re.finditer(r"@frappe\.whitelist\s*\(", text)]
        for offset in whitelist_lines:
            next_top_level = re.search(r"\n(?=def\s+\w+\()", text[offset + 1 :])
            block_end = offset + 1 + next_top_level.start() if next_top_level else len(text)
            block = text[offset:block_end]
            for match in re.finditer(r"ignore_permissions\s*=\s*True", block):
                hits.append({"file": str(path), "offset": offset + match.start()})
    return runner.check("No ignore_permissions=True in new whitelisted methods", not hits, {"hits": hits})


def gate_tree_navigation_contract(runner: GateRunner):
    customers = frappe.get_all(
        LOCATION_DOCTYPE,
        filters={"location_type": "Customer"},
        fields=["name", "location_code", "location_name", "full_path", "lft", "rgt"],
        order_by="lft asc",
    )
    samples = []
    failures = []
    for customer in customers:
        sites = frappe.get_all(
            LOCATION_DOCTYPE,
            filters={"parent_por_physical_location": customer.name, "location_type": "Site"},
            fields=["name", "location_code", "location_name", "full_path"],
            order_by="lft asc",
        )
        if not sites:
            failures.append({"customer": customer.name, "issue": "no child sites"})
            continue
        site = sites[0]
        lanes = frappe.get_all(
            LOCATION_DOCTYPE,
            filters={"parent_por_physical_location": site.name, "location_type": "Lane"},
            fields=["name", "location_code", "location_name", "full_path"],
            order_by="lft asc",
        )
        cells = []
        if lanes:
            cells = frappe.get_all(
                LOCATION_DOCTYPE,
                filters={"parent_por_physical_location": lanes[0].name, "location_type": "Cell"},
                fields=["name", "location_code", "location_name", "full_path"],
                order_by="lft asc",
            )
        if not lanes or not cells:
            failures.append({"customer": customer.name, "site": site.name, "issue": "missing lane/cell descendants"})
        samples.append({"customer": customer, "site": site, "lane": lanes[0] if lanes else None, "cell": cells[0] if cells else None})
    return runner.check(
        "Tree navigation contract renders Customer -> Site -> Lane -> Cell data",
        bool(customers) and not failures,
        {"customer_count": len(customers), "failures": failures, "samples": samples[:5]},
    )


def gate_list_readability(runner: GateRunner):
    failures = []
    meta_details = {}
    for doctype in [REQUEST_DOCTYPE, FIELD_CHANGE_DOCTYPE]:
        meta = frappe.get_meta(doctype)
        required = ["instance", "location_label", "customer", "status"]
        meta_details[doctype] = {
            field: {
                "exists": bool(meta.get_field(field)),
                "in_list_view": getattr(meta.get_field(field), "in_list_view", None) if meta.get_field(field) else None,
                "read_only": getattr(meta.get_field(field), "read_only", None) if meta.get_field(field) else None,
                "fetch_from": getattr(meta.get_field(field), "fetch_from", None) if meta.get_field(field) else None,
            }
            for field in required
        }
        for field, info in meta_details[doctype].items():
            if not info["exists"] or not info["in_list_view"]:
                failures.append({"doctype": doctype, "field": field, "issue": "missing or not list-visible", "info": info})

    request_rows = frappe.get_all(
        REQUEST_DOCTYPE,
        fields=["name", "instance", "machine_identifier", "location_label", "customer", "status"],
        limit=500,
    )
    for row in request_rows:
        if row.instance:
            inst = frappe.db.get_value(INSTANCE_DOCTYPE, row.instance, ["deployment_site", "customer"], as_dict=True)
            if inst and row.location_label != inst.deployment_site:
                failures.append({"doctype": REQUEST_DOCTYPE, "name": row.name, "issue": "location_label mismatch", "expected": inst.deployment_site, "actual": row.location_label})
            if inst and row.customer != inst.customer:
                failures.append({"doctype": REQUEST_DOCTYPE, "name": row.name, "issue": "customer mismatch", "expected": inst.customer, "actual": row.customer})
        elif row.machine_identifier and row.location_label != row.machine_identifier:
            failures.append({"doctype": REQUEST_DOCTYPE, "name": row.name, "issue": "missing machine_identifier fallback", "expected": row.machine_identifier, "actual": row.location_label})

    fc_rows = frappe.get_all(
        FIELD_CHANGE_DOCTYPE,
        fields=["name", "instance", "location_label", "customer", "status"],
        limit=500,
    )
    for row in fc_rows:
        inst = frappe.db.get_value(INSTANCE_DOCTYPE, row.instance, ["deployment_site", "customer"], as_dict=True)
        if inst and row.location_label != inst.deployment_site:
            failures.append({"doctype": FIELD_CHANGE_DOCTYPE, "name": row.name, "issue": "location_label mismatch", "expected": inst.deployment_site, "actual": row.location_label})
        if inst and row.customer != inst.customer:
            failures.append({"doctype": FIELD_CHANGE_DOCTYPE, "name": row.name, "issue": "customer mismatch", "expected": inst.customer, "actual": row.customer})

    return runner.check(
        "Field Change list readability fields are visible and populated",
        not failures,
        {"meta": meta_details, "request_rows": len(request_rows), "field_change_rows": len(fc_rows), "failures": failures[:50]},
    )


def gate_fco_register_and_importer(runner: GateRunner, seed_dir: Path, evidence_dir: Path):
    import inductone_tools.field_change as field_change

    expected_columns = [
        "fco_number",
        "date_raised",
        "requester",
        "intake_ref",
        "customer_project",
        "serial_or_location",
        "change_summary",
        "triage_outcome",
        "reference",
        "safety_regulatory",
        "disposition",
        "disposition_date",
        "implemented_date",
        "as_maintained_updated",
        "post_change_test",
        "status",
        "closed_date",
        "notes",
    ]

    original_user = frappe.session.user
    frappe.set_user("Administrator")
    failures = []
    rows = field_change.render_fco_register()
    actual_columns = list(rows[0].keys()) if rows else []
    if actual_columns != expected_columns:
        failures.append({"issue": "register column contract mismatch", "expected": expected_columns, "actual": actual_columns})

    private_files = Path(frappe.get_site_path("private", "files"))
    private_files.mkdir(parents=True, exist_ok=True)
    source = seed_dir / "fco_jotform_export.xlsx"
    target = private_files / "fco_jotform_import_validation.xlsx"
    shutil.copyfile(source, target)
    file_url = "/private/files/fco_jotform_import_validation.xlsx"
    if not frappe.db.exists("File", {"file_url": file_url}):
        file_doc = frappe.get_doc(
            {
                "doctype": "File",
                "file_name": target.name,
                "file_url": file_url,
                "is_private": 1,
            }
        )
        file_doc.insert(ignore_permissions=True)

    importer_result = field_change.import_jotform_export(file_url)
    if importer_result.get("created"):
        failures.append({"issue": "importer created rows during idempotency validation", "created": importer_result.get("created")})
    if importer_result.get("rows_read") != 19:
        failures.append({"issue": "importer row count mismatch", "result": importer_result})
    frappe.set_user(original_user)

    return runner.check(
        "SUP-FCO-R01 register export and JotForm importer contract",
        bool(rows) and not failures,
        {
            "register_rows": len(rows),
            "columns": actual_columns,
            "importer_result": importer_result,
            "failures": failures,
        },
    )


def run(args):
    frappe.init(site=args.site, sites_path=args.sites_path)
    frappe.connect()
    try:
        runner = GateRunner()
        gate_location_tree(runner, args.seed_dir)
        gate_instance_locations(runner, args.seed_dir)
        gate_field_changes_resolve(runner)
        gate_fco_map_represented(runner, args.seed_dir)
        gate_as_installed_query(runner)
        gate_reassignment_mechanism(runner)
        gate_external_builder_denial(runner)
        gate_ignore_permissions_clean(runner, args.app_path)
        gate_tree_navigation_contract(runner)
        gate_list_readability(runner)
        gate_fco_register_and_importer(runner, args.seed_dir, args.evidence_dir)
        payload = {
            "site": args.site,
            "generated_at_utc": datetime.utcnow().isoformat(timespec="seconds") + "Z",
            "ok": runner.ok,
            "gates": runner.gates,
        }
        return payload
    finally:
        frappe.destroy()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--site", required=True)
    parser.add_argument("--sites-path", required=True)
    parser.add_argument("--seed-dir", type=Path, default=Path(__file__).resolve().parent / "seeds")
    parser.add_argument("--app-path", type=Path, default=Path.cwd() / "apps" / "inductone_tools")
    parser.add_argument("--evidence-dir", type=Path, default=Path("/mnt/c/hub/frappe-sandbox/validation-evidence"))
    args = parser.parse_args()
    args.evidence_dir.mkdir(parents=True, exist_ok=True)

    payload = run(args)
    path = args.evidence_dir / "fco_as_installed_validation.json"
    path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    print(f"Evidence: {path}")
    return 0 if payload["ok"] else 1


if __name__ == "__main__":
    sys.exit(main())
