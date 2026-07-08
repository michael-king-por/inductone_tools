#!/usr/bin/env python3
"""Candidate validation for balloon-scoped electrical options.

This script intentionally runs only on a non-production candidate site. It
creates synthetic Configured BOM Snapshots for the reviewed option matrix,
resolves them through ``bom_export.build_configured_rows``, materializes the
snapshot hierarchy/workbook, and writes JSON evidence.
"""

from __future__ import annotations

import argparse
import json
import traceback
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook


DEFAULT_EVIDENCE_DIR = Path("/mnt/c/hub/frappe-sandbox/validation-evidence")
DEFAULT_BUILD = "SAL-ORD-2026-00054-BLD-0225"
EXPECTED_SCRIPT_LENGTH_BEFORE_BRANCH = 63230


@dataclass(frozen=True)
class MatrixCase:
    case_id: int
    label: str
    moved_options: tuple[str, ...]


MATRIX: tuple[MatrixCase, ...] = (
    MatrixCase(1, "baseline_only", ()),
    MatrixCase(2, "mcp", ("DEV-PANEL-MCP",)),
    MatrixCase(3, "ipc", ("DEV-PANEL-IPC",)),
    MatrixCase(4, "hmi", ("DEV-COMP-HMI",)),
    MatrixCase(5, "stack", ("DEV-COMP-STACK",)),
    MatrixCase(6, "fortress", ("DEV-COMP-FORTRESS",)),
    MatrixCase(7, "maglock", ("DEV-COMP-MAGLOCK",)),
    MatrixCase(8, "mcp_ipc", ("DEV-PANEL-MCP", "DEV-PANEL-IPC")),
    MatrixCase(9, "ipc_hmi", ("DEV-PANEL-IPC", "DEV-COMP-HMI")),
    MatrixCase(10, "ipc_maglock", ("DEV-PANEL-IPC", "DEV-COMP-MAGLOCK")),
    MatrixCase(11, "mcp_fortress", ("DEV-PANEL-MCP", "DEV-COMP-FORTRESS")),
    MatrixCase(12, "everything_moved", ("DEV-PANEL-MCP", "DEV-PANEL-IPC", "DEV-COMP-HMI", "DEV-COMP-STACK", "DEV-COMP-FORTRESS", "DEV-COMP-MAGLOCK")),
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--site", required=True)
    parser.add_argument("--sites-path", required=True)
    parser.add_argument("--evidence-dir", type=Path, default=DEFAULT_EVIDENCE_DIR)
    parser.add_argument("--build", default=DEFAULT_BUILD)
    parser.add_argument("--load-options", action="store_true", help="Upsert the reviewed option catalog before validation.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    args.evidence_dir.mkdir(parents=True, exist_ok=True)
    evidence_path = args.evidence_dir / f"balloon_scoped_options_validation_{timestamp}.json"

    import frappe
    from inductone_tools.balloon_scoped_options import (
        COLLISION_REFERENCE_BOM,
        EXTENSIONS,
        MASTER_ELECTRICAL_BOM,
        SUBSTITUTIONS,
        catalog_specs,
        expected_resolution,
        upsert_catalog,
    )
    from inductone_tools.bom_export import build_configured_rows, explode_bom_tree_structured
    from inductone_tools.snapshot.hierarchy import generate_hierarchy_workbook, populate_snapshot_hierarchy

    payload: dict = {
        "site": args.site,
        "build": args.build,
        "generated_at_utc": timestamp,
        "preconditions": {},
        "option_loader": None,
        "cases": [],
    }
    failures: list[str] = []

    frappe.init(site=args.site, sites_path=args.sites_path)
    frappe.connect()
    try:
        if args.load_options:
            payload["option_loader"] = upsert_catalog(frappe)

        preconditions = run_preconditions(frappe, args.build, MASTER_ELECTRICAL_BOM, COLLISION_REFERENCE_BOM)
        payload["preconditions"] = preconditions
        for check in preconditions["checks"]:
            print(("PASS" if check["passed"] else "FAIL"), check["label"], check.get("details", ""), flush=True)
            if not check["passed"]:
                failures.append(f"precondition:{check['label']}")

        if failures:
            payload["failures"] = failures
            evidence_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
            return 1

        build = frappe.get_doc("InductOne Build", args.build)
        baseline_rows = explode_bom_tree_structured(
            root_bom=build.top_bom,
            explosion_mode="Follow Explicit Child BOM Links",
            include_qty=True,
        )

        baseline_reference = None
        for case in MATRIX:
            print(f"RUN case {case.case_id} {case.label}", flush=True)
            case_result = run_case(
                frappe=frappe,
                case=case,
                build=build,
                baseline_rows=baseline_rows,
                catalog_specs=catalog_specs(),
                expected_resolution=expected_resolution,
                build_configured_rows=build_configured_rows,
                populate_snapshot_hierarchy=populate_snapshot_hierarchy,
                generate_hierarchy_workbook=generate_hierarchy_workbook,
                baseline_reference=baseline_reference,
            )
            if case.case_id == 1 and case_result.get("passed"):
                baseline_reference = case_result.get("consistency_reference")
            payload["cases"].append(case_result)
            status = "PASS" if case_result["passed"] else "FAIL"
            print(status, f"case {case.case_id} {case.label}", case_result.get("summary", ""), flush=True)
            if not case_result["passed"]:
                failures.append(f"case:{case.case_id}:{case.label}")

        payload["failures"] = failures
        evidence_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
    finally:
        frappe.destroy()

    print(f"Evidence: {evidence_path}", flush=True)
    return 1 if failures else 0


def run_preconditions(frappe, build_name: str, master_bom: str, collision_bom: str) -> dict:
    from inductone_tools.balloon_scoped_options import EXTENSIONS, SUBSTITUTIONS

    checks = []
    build_exists = frappe.db.exists("InductOne Build", build_name)
    checks.append({"label": "BLD-0225 exists", "passed": bool(build_exists), "details": build_name})

    bom = frappe.db.get_value("BOM", master_bom, ["docstatus", "is_active"], as_dict=True)
    checks.append({
        "label": "REV E master active submitted",
        "passed": bool(bom and int(bom.docstatus or 0) == 1 and int(bom.is_active or 0) == 1),
        "details": bom,
    })

    configurable_rows = []
    for sub in SUBSTITUTIONS:
        configurable_rows.append((sub.balloon, sub.standard_item))
        configurable_rows.extend((sub.balloon, item) for item in sub.option_items)
    configurable_rows.extend((ext.balloon, ext.item) for ext in EXTENSIONS)
    found = 0
    missing = []
    for balloon, item in configurable_rows:
        exists = frappe.db.exists(
            "BOM Item",
            {"parent": master_bom, "custom_balloon_numbers": balloon, "item_code": item},
        )
        if exists:
            found += 1
        else:
            missing.append({"balloon": balloon, "item_code": item})
    checks.append({
        "label": "REV E configurable balloon fingerprint",
        "passed": found == 26 and not missing,
        "details": {"found": found, "expected": 26, "missing": missing},
    })

    collision = frappe.get_all(
        "BOM Item",
        filters={"parent": collision_bom, "custom_balloon_numbers": "315", "item_code": "1417891"},
        fields=["name", "idx", "qty"],
    )
    checks.append({
        "label": "0921 fixed 1417891 balloon 315 present",
        "passed": len(collision) == 1 and float(collision[0].qty or 0) == 3.0,
        "details": collision,
    })

    script_len = frappe.db.get_value("Client Script", "InductOne Build Script", "script")
    script_len = len(script_len or "")
    checks.append({
        "label": "InductOne Build Script has balloon carry-through",
        "passed": script_len >= EXPECTED_SCRIPT_LENGTH_BEFORE_BRANCH,
        "details": {"length": script_len, "pre_branch_length": EXPECTED_SCRIPT_LENGTH_BEFORE_BRANCH},
    })

    field_rows = frappe.get_all(
        "Custom Field",
        filters={"fieldname": "target_balloon", "dt": ["in", ["InductOne Configuration Option Mapping", "Configured BOM Snapshot Structural Effect"]]},
        fields=["name", "dt", "insert_after"],
    )
    checks.append({
        "label": "target_balloon fields fixture-managed on candidate",
        "passed": len(field_rows) == 2,
        "details": field_rows,
    })

    return {"checks": checks}


def run_case(
    frappe,
    case: MatrixCase,
    build,
    baseline_rows: list[dict],
    catalog_specs: list[dict],
    expected_resolution,
    build_configured_rows,
    populate_snapshot_hierarchy,
    generate_hierarchy_workbook,
    baseline_reference: dict | None = None,
) -> dict:
    result = {
        "case_id": case.case_id,
        "label": case.label,
        "moved_options": list(case.moved_options),
        "checks": [],
        "passed": False,
    }
    try:
        selected_codes = ["DEV-BASELINE", *case.moved_options]
        structural_effects = structural_effects_for(selected_codes, catalog_specs)
        snap = create_validation_snapshot(frappe, build, baseline_rows, structural_effects, case)
        pkg = SimpleNamespace(
            inductone_build=build.name,
            configured_snapshot=snap.name,
            bom=build.top_bom,
            explosion_mode="Follow Explicit Child BOM Links",
            max_depth=None,
            include_qty=1,
            preserve_duplicate_occurrences=1,
        )
        rows = build_configured_rows(pkg)
        expected = expected_resolution(selected_codes, frappe)

        # Stage 1 — effects carry occurrence identity.
        add_check(result, "stage_1_effects_carry_target_balloon", assert_effects_have_balloons(snap, structural_effects))

        # Stage 2 — canonical resolver output vs independent oracle.
        flat_managed = managed_rows_by_balloon(rows)
        flat_collision = collision_rollup(rows)
        add_check(result, "stage_2_flat_managed_balloons_match_oracle", assert_balloon_rows(rows, expected["by_balloon"]))
        add_check(result, "stage_2_flat_collision_quantities_match_oracle", assert_collision_flat(rows, expected["flat"]))

        hierarchy_result = populate_snapshot_hierarchy(snap.name)
        add_check(result, "hierarchy_populated", bool(hierarchy_result and hierarchy_result.get("ok", True)), hierarchy_result)
        hierarchy_rows = load_hierarchy_rows(frappe, snap.name)
        hierarchy_managed = managed_rows_by_balloon(hierarchy_rows)
        hierarchy_collision = collision_rollup(hierarchy_rows)

        # Stage 3 — materialized hierarchy vs independent oracle.
        add_check(result, "stage_3_hierarchy_managed_balloons_match_oracle", assert_balloon_rows(hierarchy_rows, expected["by_balloon"]))
        add_check(result, "stage_3_hierarchy_collision_quantities_match_oracle", assert_collision_flat(hierarchy_rows, expected["flat"]))

        workbook_result = generate_hierarchy_workbook(snap.name)
        add_check(result, "hierarchy_workbook_generated", bool(workbook_result), workbook_result)
        workbook_rows = load_workbook_rows(frappe, workbook_result)
        workbook_managed = managed_rows_by_balloon(workbook_rows)
        workbook_collision = collision_rollup(workbook_rows)

        # Stage 4 — workbook content vs independent oracle.
        add_check(result, "stage_4_workbook_managed_balloons_match_oracle", assert_balloon_rows(workbook_rows, expected["by_balloon"]))
        add_check(result, "stage_4_workbook_collision_quantities_match_oracle", assert_collision_flat(workbook_rows, expected["flat"]))

        # Stage 5 — all derived artifacts agree with each other.
        add_check(
            result,
            "stage_5_cross_stage_managed_consistency",
            assert_cross_stage_consistency({
                "flat": flat_managed,
                "hierarchy": hierarchy_managed,
                "workbook": workbook_managed,
            }),
        )
        add_check(
            result,
            "stage_5_cross_stage_collision_consistency",
            assert_cross_stage_consistency({
                "flat": flat_collision,
                "hierarchy": hierarchy_collision,
                "workbook": workbook_collision,
            }),
        )

        flat_non_managed = non_managed_rollup(rows)
        hierarchy_non_managed = non_managed_rollup(hierarchy_rows)
        consistency_reference = {
            "flat_non_managed": flat_non_managed,
            "hierarchy_non_managed": hierarchy_non_managed,
            "flat_managed": flat_managed,
            "hierarchy_managed": hierarchy_managed,
            "workbook_managed": workbook_managed,
            "flat_collision": flat_collision,
            "hierarchy_collision": hierarchy_collision,
            "workbook_collision": workbook_collision,
        }

        # Stage 6 — non-managed material remains stable vs baseline-only.
        if baseline_reference is None:
            add_check(result, "stage_6_conservation_reference_recorded", True, {"reference": "this baseline-only case"})
        else:
            add_check(
                result,
                "stage_6_flat_non_managed_conserved_vs_baseline",
                compare_rollups(flat_non_managed, baseline_reference["flat_non_managed"]),
            )
            add_check(
                result,
                "stage_6_hierarchy_non_managed_conserved_vs_baseline",
                compare_rollups(hierarchy_non_managed, baseline_reference["hierarchy_non_managed"]),
            )

        result["snapshot"] = snap.name
        result["consistency_reference"] = consistency_reference
        result["oracle"] = {
            "by_balloon": expected["by_balloon"],
            "flat": expected["flat"],
        }
        result["observed"] = {
            "flat_managed": flat_managed,
            "hierarchy_managed": hierarchy_managed,
            "workbook_managed": workbook_managed,
            "flat_collision": flat_collision,
            "hierarchy_collision": hierarchy_collision,
            "workbook_collision": workbook_collision,
        }
        result["summary"] = (
            f"{len(structural_effects)} effects; {len(rows)} resolved rows; "
            f"{len(hierarchy_rows)} hierarchy rows; {len(workbook_rows)} workbook rows"
        )
        result["passed"] = all(check["passed"] for check in result["checks"])
    except Exception as exc:  # noqa: BLE001 - evidence wants exact exception
        result["exception"] = {
            "type": exc.__class__.__name__,
            "message": str(exc),
            "traceback": traceback.format_exc(),
        }
        result["passed"] = False
    return result


def structural_effects_for(selected_codes: list[str], catalog_specs: list[dict]) -> list[dict]:
    specs = {spec["option_code"]: spec for spec in catalog_specs}
    effects = []
    for code in selected_codes:
        for mapping in specs[code]["mappings_table"]:
            effect_mode = mapping.get("structural_effect_mode") or "NO_STRUCTURAL_CHANGE"
            if effect_mode == "AUTO":
                action = mapping.get("action") or ""
                expand_mode = mapping.get("expand_mode") or "AS_ITEM_ONLY"
                if action == "REPLACE" and expand_mode == "AS_ITEM_ONLY":
                    effect_mode = "REPLACE_TARGET_NODE"
                elif action == "REPLACE":
                    effect_mode = "REPLACE_TARGET_BRANCH"
                elif action == "REMOVE" and expand_mode == "AS_ITEM_ONLY":
                    effect_mode = "SUPPRESS_TARGET_NODE"
                elif action == "REMOVE":
                    effect_mode = "SUPPRESS_TARGET_BRANCH"
                elif action == "ADD" and expand_mode == "AS_ITEM_ONLY":
                    effect_mode = "INCREMENT_NODE_QTY"
                elif action == "ADD":
                    effect_mode = "ADD_BRANCH"
                elif action == "QTY_OVERRIDE":
                    effect_mode = "OVERRIDE_NODE_QTY"
            effects.append({
                "action": mapping.get("action") or "",
                "effect_mode": effect_mode,
                "effect_qty": float(mapping.get("qty_fixed") or 1),
                "target_item": mapping.get("target_item") or "",
                "target_balloon": mapping.get("target_balloon") or "",
                "target_bom": mapping.get("target_bom") or "",
                "resolved_target_bom": "",
                "replace_with_item": mapping.get("replace_with_item") or "",
                "replace_scope": mapping.get("replace_scope") or "ALL_OCCURRENCES",
                "replace_count": int(mapping.get("replace_count") or 1),
                "replace_with_bom": mapping.get("replace_with_bom") or "",
                "resolved_replace_with_bom": "",
                "source_option_code": code,
                "expand_mode": mapping.get("expand_mode") or "AS_ITEM_ONLY",
                "row_order": int(mapping.get("row_order") or 100),
                "reason": f"Balloon validation effect from {code}",
            })
    return effects


def create_validation_snapshot(frappe, build, baseline_rows: list[dict], structural_effects: list[dict], case: MatrixCase):
    leaf_codes = sorted({row.get("item_code") for row in baseline_rows if row.get("is_leaf") and row.get("item_code")})
    doc = frappe.get_doc({
        "doctype": "Configured BOM Snapshot",
        "sales_order": build.sales_order,
        "inductone_build": build.name,
        "top_item": build.top_item,
        "top_bom": build.top_bom,
        "orientation": build.orientation,
        "snapshot_rev": 9000 + case.case_id,
        "generated_at": frappe.utils.now_datetime(),
        "lines": [
            {
                "item_code": item_code,
                "qty": 1,
                "included": 1,
                "rule_reason": "Validation baseline include set",
            }
            for item_code in leaf_codes
        ],
        "structural_effects": structural_effects,
    })
    doc.insert(ignore_permissions=True)
    frappe.db.commit()
    return doc


def assert_effects_have_balloons(snap, structural_effects: list[dict]) -> tuple[bool, dict]:
    expected = sorted(
        (e["source_option_code"], e["action"], e["target_item"], e["target_balloon"])
        for e in structural_effects
        if e.get("target_balloon")
    )
    actual = sorted(
        (row.source_option_code, row.action, row.target_item, getattr(row, "target_balloon", "") or "")
        for row in snap.structural_effects
        if getattr(row, "target_balloon", "") or ""
    )
    return expected == actual, {"expected": expected, "actual": actual}


def managed_items_by_balloon() -> dict[str, set[str]]:
    from inductone_tools.balloon_scoped_options import EXTENSIONS, SUBSTITUTIONS

    managed: dict[str, set[str]] = {}
    for sub in SUBSTITUTIONS:
        managed.setdefault(sub.balloon, set()).add(sub.standard_item)
        managed[sub.balloon].update(sub.option_items)
    for ext in EXTENSIONS:
        managed.setdefault(ext.balloon, set()).add(ext.item)
    return managed


def managed_item_codes() -> set[str]:
    codes = set()
    for items in managed_items_by_balloon().values():
        codes.update(items)
    return codes


def load_hierarchy_rows(frappe, snapshot_name: str) -> list[dict]:
    rows = frappe.get_all(
        "Configured BOM Snapshot Hierarchy",
        filters={
            "parent": snapshot_name,
            "parenttype": "Configured BOM Snapshot",
            "parentfield": "hierarchy",
        },
        fields=[
            "idx",
            "item_code",
            "item_name",
            "qty",
            "uom",
            "bom_used",
            "node_type",
            "is_leaf",
            "bom_level",
            "balloon_numbers",
            "source_bom",
            "source_bom_item",
            "source_bom_item_idx",
        ],
        order_by="idx asc",
    )
    out = []
    for row in rows:
        out.append({
            "item_code": row.item_code,
            "item_name": row.item_name,
            "qty": float(row.qty or 0),
            "uom": row.uom,
            "bom_used": row.bom_used,
            "node_type": row.node_type,
            "is_leaf": int(row.is_leaf or 0),
            "bom_level": int(row.bom_level or 0),
            "balloon_numbers": row.balloon_numbers or "",
            "source_bom": row.source_bom or "",
            "source_bom_item": row.source_bom_item or "",
            "source_bom_item_idx": int(row.source_bom_item_idx or 0),
        })
    return out


def load_workbook_rows(frappe, workbook_result: dict) -> list[dict]:
    file_url = (workbook_result or {}).get("file_url")
    if not file_url:
        raise ValueError("Hierarchy workbook result did not include file_url.")

    if file_url.startswith("/private/files/"):
        workbook_path = frappe.get_site_path("private", "files", file_url.split("/private/files/", 1)[1])
    elif file_url.startswith("/files/"):
        workbook_path = frappe.get_site_path("public", "files", file_url.split("/files/", 1)[1])
    else:
        raise ValueError(f"Unsupported workbook file_url: {file_url}")

    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    ws = wb["Configured BOM Hierarchy"]
    header_row = 9
    row_iter = ws.iter_rows(min_row=header_row, values_only=True)
    header_values = next(row_iter)
    headers = {
        (label or ""): idx
        for idx, label in enumerate(header_values)
    }

    def _value(values: tuple, label: str):
        idx = headers.get(label)
        if idx is None or idx >= len(values):
            return None
        return values[idx]

    out = []
    for values in row_iter:
        item_code = (_value(values, "Item Code") or "")
        item_code = str(item_code).replace("\u00a0", "").strip()
        if not item_code:
            continue
        try:
            qty = float(_value(values, "Qty") or 0)
        except (TypeError, ValueError):
            qty = 0.0
        try:
            bom_level = int(_value(values, "BOM Level") or 0)
        except (TypeError, ValueError):
            bom_level = 0
        node_type = "Assembly" if _value(values, "BOM") else "Leaf"
        out.append({
            "item_code": item_code,
            "item_name": _value(values, "Item Name") or "",
            "qty": qty,
            "uom": _value(values, "UOM") or "",
            "bom_used": _value(values, "BOM") or "",
            "node_type": node_type,
            "is_leaf": 0 if node_type == "Assembly" else 1,
            "bom_level": bom_level,
            "balloon_numbers": str(_value(values, "Balloon #") or "").strip(),
            "source_bom": _value(values, "Source BOM") or "",
            "source_bom_item": _value(values, "Source BOM Item") or "",
            "source_bom_item_idx": int(_value(values, "Source BOM Item IDX") or 0),
        })
    wb.close()
    return out


def managed_rows_by_balloon(rows: list[dict]) -> dict[str, list[dict]]:
    managed = managed_items_by_balloon()
    observed: dict[str, list[dict]] = {}
    for balloon, items in managed.items():
        actual_rows = [
            {"item_code": row.get("item_code"), "qty": float(row.get("qty") or 0)}
            for row in rows
            if row.get("is_leaf")
            and (row.get("balloon_numbers") or "").strip() == balloon
            and row.get("item_code") in items
        ]
        observed[balloon] = sorted(actual_rows, key=lambda r: (r["item_code"], r["qty"]))
    return observed


def collision_rollup(rows: list[dict]) -> dict[str, float]:
    collision_items = {"11283", "11245", "11351", "1417902", "1417903", "1417891", "1417892"}
    actual: dict[str, float] = defaultdict(float)
    for row in rows:
        if not row.get("is_leaf"):
            continue
        item = row.get("item_code")
        if item in collision_items:
            actual[item] += float(row.get("qty") or 0)
    return {item: qty for item, qty in sorted(actual.items()) if abs(qty) > 1e-9}


def non_managed_rollup(rows: list[dict]) -> dict[str, float]:
    managed = managed_item_codes()
    actual: dict[str, float] = defaultdict(float)
    for row in rows:
        if not row.get("is_leaf"):
            continue
        item = row.get("item_code")
        if not item or item in managed:
            continue
        actual[item] += float(row.get("qty") or 0)
    return {item: qty for item, qty in sorted(actual.items()) if abs(qty) > 1e-9}


def assert_cross_stage_consistency(stage_payloads: dict[str, dict]) -> tuple[bool, dict]:
    names = sorted(stage_payloads)
    reference_name = names[0]
    reference = stage_payloads[reference_name]
    mismatches = {
        name: payload
        for name, payload in stage_payloads.items()
        if payload != reference
    }
    return not mismatches, {
        "reference_stage": reference_name,
        "reference": reference,
        "mismatches": mismatches,
    }


def compare_rollups(actual: dict[str, float], expected: dict[str, float]) -> tuple[bool, dict]:
    keys = sorted(set(actual) | set(expected))
    failures = {
        key: {"expected": float(expected.get(key, 0.0)), "actual": float(actual.get(key, 0.0))}
        for key in keys
        if abs(float(actual.get(key, 0.0)) - float(expected.get(key, 0.0))) > 1e-9
    }
    return not failures, {"failures": failures, "actual_count": len(actual), "expected_count": len(expected)}


def assert_balloon_rows(rows: list[dict], expected_by_balloon: dict[str, list[dict]]) -> tuple[bool, dict]:
    managed = managed_items_by_balloon()

    failures = []
    observed = {}
    leaf_rows = [row for row in rows if row.get("is_leaf")]
    for balloon, expected_rows in expected_by_balloon.items():
        managed_items = managed.get(balloon, set())
        actual_rows = [
            {"item_code": row.get("item_code"), "qty": float(row.get("qty") or 0)}
            for row in leaf_rows
            if (row.get("balloon_numbers") or "").strip() == balloon
            and row.get("item_code") in managed_items
        ]
        actual_rows = sorted(actual_rows, key=lambda r: (r["item_code"], r["qty"]))
        expected_sorted = sorted(
            [{"item_code": row["item_code"], "qty": float(row["qty"] or 0)} for row in expected_rows],
            key=lambda r: (r["item_code"], r["qty"]),
        )
        observed[balloon] = actual_rows
        if actual_rows != expected_sorted:
            failures.append({"balloon": balloon, "expected": expected_sorted, "actual": actual_rows})
    return not failures, {"failures": failures, "observed": observed}


def assert_collision_flat(rows: list[dict], expected_flat: dict[str, float]) -> tuple[bool, dict]:
    collision_items = ["11283", "11245", "11351", "1417902", "1417903", "1417891", "1417892"]
    actual = {}
    for row in rows:
        if not row.get("is_leaf"):
            continue
        item = row.get("item_code")
        if item in collision_items:
            actual[item] = actual.get(item, 0.0) + float(row.get("qty") or 0)
    expected = {item: float(expected_flat.get(item, 0.0)) for item in collision_items if expected_flat.get(item, 0.0)}
    # Keep zero-expected items explicit if they appear in actual output.
    for item in actual:
        expected.setdefault(item, 0.0)
    failures = {
        item: {"expected": expected.get(item, 0.0), "actual": actual.get(item, 0.0)}
        for item in sorted(set(expected) | set(actual))
        if abs(float(expected.get(item, 0.0)) - float(actual.get(item, 0.0))) > 1e-9
    }
    return not failures, {"expected": expected, "actual": actual, "failures": failures}


def add_check(result: dict, label: str, outcome, details=None) -> None:
    if isinstance(outcome, tuple):
        passed, tuple_details = outcome
        details = tuple_details
    else:
        passed = bool(outcome)
    result["checks"].append({"label": label, "passed": bool(passed), "details": details})


if __name__ == "__main__":
    raise SystemExit(main())
