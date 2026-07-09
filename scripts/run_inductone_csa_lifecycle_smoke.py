#!/usr/bin/env python3
"""Candidate-only InductOne CSA lifecycle smoke test.

This script creates a synthetic candidate build cloned from a known restored
production build, then drives the real server-side workflow functions:

1. create Build / Snapshot / Configuration Order / BOM Export Package
2. approve the required top-BOM engineering signoff if needed
3. allocate a system serial from the candidate builder tranche
4. release to builder, including generated release manifest + serial workbook
5. acknowledge builder release
6. upload a filled builder completion workbook through the upload parser
7. mark completion Reviewed
8. accept completion, creating locked As-Built and InductOne Instance

It intentionally mutates only the candidate sandbox. Do not run this against
production unless the system owner explicitly wants real validation records.
"""

from __future__ import annotations

import argparse
import io
import json
import sys
import traceback
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace

from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from run_balloon_scoped_options_validation import (  # noqa: E402
    create_validation_snapshot,
    run_preconditions,
    structural_effects_for,
)
from run_per_option_snapshot_diff_reports import (  # noqa: E402
    DEFAULT_STABLE_OPTION_CODES,
    SnapshotCase,
    load_site_catalog_specs,
    resolve_option_codes,
    selected_codes_for,
)


DEFAULT_SITE = "inductone-candidate.localhost"
DEFAULT_SITES_PATH = "/home/michaelplusone/frappe-sandbox/benches/candidate-bench/sites"
DEFAULT_EVIDENCE_DIR = Path("/mnt/c/hub/frappe-sandbox/validation-evidence")
DEFAULT_SOURCE_BUILD = "SAL-ORD-2026-00054-BLD-0225"
DEFAULT_MANAGER_USER = "christina.gt@plusonerobotics.com"
DEFAULT_ENGINEERING_USER = "shaun.edwards@plusonerobotics.com"
DEFAULT_BUILDER_USER = "motion.builder@plusonerobotics.com"
EXPECTED_SITE_FRAGMENT = "candidate"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run candidate-only InductOne CSA lifecycle smoke.")
    parser.add_argument("--site", default=DEFAULT_SITE)
    parser.add_argument("--sites-path", default=DEFAULT_SITES_PATH)
    parser.add_argument("--evidence-dir", type=Path, default=DEFAULT_EVIDENCE_DIR)
    parser.add_argument("--source-build", default=DEFAULT_SOURCE_BUILD)
    parser.add_argument("--manager-user", default=DEFAULT_MANAGER_USER)
    parser.add_argument("--engineering-user", default=DEFAULT_ENGINEERING_USER)
    parser.add_argument("--builder-user", default=DEFAULT_BUILDER_USER)
    parser.add_argument(
        "--confirm-candidate",
        action="store_true",
        help="Required guard acknowledging this script mutates only the candidate sandbox.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.confirm_candidate or EXPECTED_SITE_FRAGMENT not in args.site:
        print(
            "Refusing to run mutating lifecycle smoke unless --confirm-candidate "
            f"is supplied and site contains '{EXPECTED_SITE_FRAGMENT}'.",
            file=sys.stderr,
        )
        return 2

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    args.evidence_dir.mkdir(parents=True, exist_ok=True)
    evidence_path = args.evidence_dir / f"inductone_csa_lifecycle_smoke_{timestamp}.json"

    import frappe

    payload: dict = {
        "site": args.site,
        "source_build": args.source_build,
        "generated_at_utc": timestamp,
        "candidate_only": True,
        "steps": [],
        "created_records": {},
        "observations": [],
        "failures": [],
    }

    frappe.init(site=args.site, sites_path=args.sites_path)
    frappe.connect()
    try:
        run_lifecycle(args, frappe, payload, timestamp)
    except Exception as exc:  # noqa: BLE001 - evidence needs exact failure
        payload["failures"].append({
            "type": exc.__class__.__name__,
            "message": str(exc),
            "traceback": traceback.format_exc(),
        })
    finally:
        payload["passed"] = not payload["failures"] and all(step.get("passed") for step in payload["steps"])
        evidence_path.write_text(json.dumps(payload, indent=2, default=str), encoding="utf-8")
        print(f"Evidence: {evidence_path}")
        print("SUMMARY", "PASS" if payload["passed"] else "FAIL")
        for step in payload["steps"]:
            print("PASS" if step.get("passed") else "FAIL", step["label"], step.get("summary", ""))
        frappe.destroy()

    return 0 if payload["passed"] else 1


def run_lifecycle(args, frappe, payload: dict, timestamp: str) -> None:
    from inductone_tools.balloon_scoped_options import (
        COLLISION_REFERENCE_BOM,
        MASTER_ELECTRICAL_BOM,
        expected_resolution,
    )
    from inductone_tools.bom_export import explode_bom_tree_structured, generate_now
    from inductone_tools.builder_release import (
        acknowledge_builder_release,
        check_builder_release_readiness,
        release_to_builder_now,
    )
    from inductone_tools.build_completion import create_completion_from_upload
    from inductone_tools.build_completion_accept import accept_completion_create_as_built
    from inductone_tools.build_completion_workbook_parser import parse_builder_workbook
    from inductone_tools.engineering_signoff import (
        approve_signoff,
        get_current_signoff_status,
        request_signoff,
    )
    from inductone_tools.serial_allocation.release import allocate_serial_for_build
    from inductone_tools.snapshot.hierarchy import generate_hierarchy_workbook, populate_snapshot_hierarchy

    # Step 0: preconditions on restored REV E data.
    preconditions = run_preconditions(frappe, args.source_build, MASTER_ELECTRICAL_BOM, COLLISION_REFERENCE_BOM)
    payload["preconditions"] = preconditions
    add_step(
        payload,
        "preconditions",
        all(check["passed"] for check in preconditions["checks"]),
        {"checks": preconditions["checks"]},
    )
    assert_last_step(payload)

    source_build = frappe.get_doc("InductOne Build", args.source_build)

    # Step 1: create synthetic build cloned from the source build.
    frappe.set_user(args.manager_user)
    build = create_smoke_build(frappe, source_build, timestamp)
    payload["created_records"]["build"] = build.name
    add_step(payload, "create_synthetic_build", True, summarize_doc(build, ["name", "sales_order", "top_item", "top_bom", "builder_supplier", "build_status"]))

    # Step 2: freeze a baseline snapshot and hierarchy workbook through existing machinery.
    stable_codes = resolve_option_codes(frappe, DEFAULT_STABLE_OPTION_CODES)
    selected_codes = resolve_option_codes(frappe, [*stable_codes, *selected_codes_for(())])
    attach_selected_options_to_build(frappe, build, selected_codes)
    baseline_rows = explode_bom_tree_structured(
        root_bom=build.top_bom,
        explosion_mode="Follow Explicit Child BOM Links",
        include_qty=True,
    )
    structural_effects = structural_effects_for(selected_codes, load_site_catalog_specs(frappe, selected_codes))
    snapshot = create_validation_snapshot(
        frappe,
        build,
        baseline_rows,
        structural_effects,
        SnapshotCase(9801, "csa_lifecycle_baseline", ()),
    )
    hierarchy_result = populate_snapshot_hierarchy(snapshot.name)
    workbook_result = generate_hierarchy_workbook(snapshot.name)
    build.latest_snapshot = snapshot.name
    build.selected_snapshot = snapshot.name
    build.snapshot_rev = int(getattr(build, "snapshot_rev", 0) or 0) + 1
    build.save(ignore_permissions=True)
    payload["created_records"]["snapshot"] = snapshot.name
    add_step(payload, "generate_snapshot_and_hierarchy", True, {
        "snapshot": snapshot.name,
        "selected_options": selected_codes,
        "structural_effect_count": len(structural_effects),
        "hierarchy_result": hierarchy_result,
        "workbook_result": workbook_result,
    })

    # Step 3: create Configuration Order tied to the current snapshot.
    co = create_configuration_order(frappe, build, snapshot.name, selected_codes)
    payload["created_records"]["configuration_order"] = co.name
    add_step(payload, "create_configuration_order", True, summarize_doc(co, ["name", "co_status", "inductone_build", "snapshot", "builder_supplier"]))

    # Step 4: generate a real but lightweight BOM Export Package.
    package = create_minimal_generated_package(frappe, build, co.name, snapshot.name, generate_now)
    payload["created_records"]["bom_export_package"] = package.name
    add_step(payload, "generate_bom_export_package", bool(package.output_zip), {
        "package": package.name,
        "status": package.status,
        "output_zip": package.output_zip,
        "result_count": len(package.results or []),
        "note": "Lifecycle smoke uses a minimal generated ZIP. Heavy PDF/DXF part-documentation payload is covered by balloon_export_zip_closeout evidence.",
    })
    assert_last_step(payload)

    # Step 5: ensure required engineering signoff for top BOM.
    top_bom_signoff_before = get_current_signoff_status("BOM", build.top_bom)
    signoff_result = None
    if top_bom_signoff_before != "Approved":
        frappe.set_user(args.engineering_user)
        signoff_request = request_signoff("BOM", build.top_bom)
        signoff_result = approve_signoff(signoff_request["signoff_name"], notes="Candidate lifecycle smoke top-BOM release gate approval.")
    top_bom_signoff_after = get_current_signoff_status("BOM", build.top_bom)
    add_step(payload, "ensure_top_bom_signoff", top_bom_signoff_after == "Approved", {
        "before": top_bom_signoff_before,
        "after": top_bom_signoff_after,
        "created_or_approved": signoff_result,
    })
    assert_last_step(payload)

    # Step 6: allocate serial using real builder tranche logic.
    frappe.set_user(args.manager_user)
    serial_result = allocate_serial_for_build(build.name)
    build.reload()
    co.reload()
    add_step(payload, "allocate_serial", bool(serial_result.get("system_serial") and build.system_serial), {
        "serial_result": serial_result,
        "build_system_serial": build.system_serial,
        "co_system_serial": getattr(co, "system_serial", None),
    })
    assert_last_step(payload)

    # Step 7: readiness must now pass.
    readiness = check_builder_release_readiness(build.name)
    add_step(payload, "release_readiness", bool(readiness.get("ready")), readiness)
    assert_last_step(payload)

    # Step 8: release to builder through real method. This generates manifest + serial workbook.
    release_result = release_to_builder_now(build.name, package_name=package.name, note="Candidate lifecycle smoke release.")
    build.reload()
    co.reload()
    add_step(payload, "release_to_builder", bool(release_result.get("ok") and co.co_status == "Released"), {
        "release_result": release_result,
        "build_status": build.build_status,
        "co_status": co.co_status,
        "document_count": len(co.documents or []),
    })
    assert_last_step(payload)

    # Step 9: acknowledge as external builder.
    frappe.set_user(args.builder_user)
    ack_result = acknowledge_builder_release(build.name, note="Candidate lifecycle smoke acknowledgement.")
    co.reload()
    add_step(payload, "builder_acknowledgement", bool(ack_result.get("ok") and co.co_status == "Awaiting Completion"), {
        "ack_result": ack_result,
        "co_status": co.co_status,
        "acknowledged_by": getattr(co, "acknowledged_by", None),
    })
    assert_last_step(payload)

    # Step 10: fill the generated builder workbook and upload through real parser flow.
    frappe.set_user(args.builder_user)
    template_url = release_result.get("serial_template_file_url")
    completion_file = create_filled_completion_workbook(frappe, template_url, build.system_serial, build.name, timestamp)
    parsed = parse_builder_workbook(read_file_url_bytes(frappe, completion_file["file_url"]))
    completion_result = create_completion_from_upload(
        build.name,
        completion_file["file_url"],
        submitted_by_name="Candidate Lifecycle Smoke",
        builder_reference=f"CSA-LIFECYCLE-{timestamp}",
        completion_notes="Synthetic candidate lifecycle completion upload.",
    )
    add_step(payload, "upload_builder_completion", bool(completion_result.get("ok")), {
        "completion_result": completion_result,
        "uploaded_file": completion_file,
        "parsed_component_rows": len(parsed["components"]),
        "filled_component_rows": sum(1 for row in parsed["components"] if row["serial_number"]),
    })
    assert_last_step(payload)

    # Step 11: review the submitted completion.
    frappe.set_user(args.manager_user)
    completion = frappe.get_doc("InductOne Build Completion", completion_result["completion_name"])
    completion.status = "Reviewed"
    if hasattr(completion, "review_notes"):
        completion.review_notes = "Candidate lifecycle smoke review passed."
    completion.save(ignore_permissions=True)
    add_step(payload, "review_completion", completion.status == "Reviewed", {
        "completion": completion.name,
        "status": completion.status,
        "serial_rows": len(completion.serials or []),
        "reviewed_by": getattr(completion, "reviewed_by", None),
        "reviewed_at": getattr(completion, "reviewed_at", None),
    })
    assert_last_step(payload)

    # Step 12: accept completion; this must create locked As-Built + Instance and close CO.
    accept_result = accept_completion_create_as_built(completion.name, as_built_notes="Candidate lifecycle smoke accepted.")
    build.reload()
    co.reload()
    as_built = frappe.get_doc("InductOne As-Built Record", accept_result["as_built_name"])
    instance = frappe.get_doc("InductOne Instance", accept_result["instance_name"])
    as_built_serial_count = len(as_built.serials or [])
    instance_serial_count = len(instance.component_serials or [])
    add_step(payload, "accept_completion_create_as_built", bool(accept_result.get("ok")) and instance_serial_count == as_built_serial_count, {
        "accept_result": accept_result,
        "build_status": build.build_status,
        "build_completion_status": build.completion_status,
        "co_status": co.co_status,
        "as_built_status": as_built.status,
        "as_built_serial_rows": as_built_serial_count,
        "instance_status": instance.status,
        "instance_system_serial": instance.system_serial,
        "instance_component_serial_rows": instance_serial_count,
        "serial_copy_check": "PASS" if instance_serial_count == as_built_serial_count else "FAIL",
    })
    assert_last_step(payload)

    payload["final_records"] = {
        "build": summarize_doc(build, ["name", "build_status", "completion_status", "system_serial", "latest_config_order", "latest_build_completion", "as_built_record"]),
        "configuration_order": summarize_doc(co, ["name", "co_status", "system_serial", "snapshot"]),
        "as_built": summarize_doc(as_built, ["name", "status", "system_serial", "inductone_build", "configuration_order", "build_completion"]),
        "instance": summarize_doc(instance, ["name", "status", "system_serial", "inductone_build", "as_built_record", "configuration_order"]),
    }
    payload["oracle_baseline"] = {
        "expected_resolution_keys": sorted(expected_resolution(selected_codes, frappe)["by_balloon"].keys()),
    }
    payload["observations"].append(
        "This smoke validates the happy path. Direct negative hardening for acknowledgement is covered by run_inductone_csa_hardening_gates.py and run_method_negative_tests.py."
    )


def create_smoke_build(frappe, source_build, timestamp: str):
    name = f"{source_build.name}-CSA-SMOKE-{timestamp.replace('T', '').replace('Z', '')}"
    doc = frappe.new_doc("InductOne Build")
    doc.name = name
    doc.sales_order = source_build.sales_order
    doc.sales_order_item_idx = source_build.sales_order_item_idx
    doc.sales_order_item_row_name = getattr(source_build, "sales_order_item_row_name", None)
    doc.customer_project_label = f"Candidate lifecycle smoke cloned from {source_build.name}"
    doc.top_item = source_build.top_item
    doc.top_bom = source_build.top_bom
    doc.orientation = source_build.orientation or "Right-Hand"
    doc.builder_supplier = source_build.builder_supplier
    doc.builder_poc = getattr(source_build, "builder_poc", None)
    doc.builder_site = getattr(source_build, "builder_site", None)
    doc.builder_po_reference = "Candidate lifecycle smoke"
    doc.build_status = "DRAFT"
    doc.completion_status = "Open"
    doc.insert(ignore_permissions=True, ignore_mandatory=True)
    frappe.db.commit()
    return doc


def attach_selected_options_to_build(frappe, build, selected_codes: list[str]) -> None:
    build.set("selections", [])
    for idx, code in enumerate(selected_codes, start=1):
        option_name = frappe.db.get_value("InductOne Configuration Option", {"option_code": code}, "name")
        option = frappe.get_doc("InductOne Configuration Option", option_name)
        build.append("selections", {
            "option": option.name,
            "option_code": option.option_code,
            "option_name": option.option_name,
            "option_category": getattr(option, "option_category", None),
            "option_group": getattr(option, "option_group", None),
            "selected": 1,
            "qty": 1,
            "sort_order": getattr(option, "sort_order", idx * 10),
        })
    build.save(ignore_permissions=True)
    frappe.db.commit()


def create_configuration_order(frappe, build, snapshot_name: str, selected_codes: list[str]):
    co = frappe.new_doc("InductOne Configuration Order")
    co.inductone_build = build.name
    co.co_status = "Draft"
    co.config_order_rev = 1
    co.generated_at = frappe.utils.now_datetime()
    co.generated_by = frappe.session.user
    co.builder_supplier = build.builder_supplier
    co.sales_order = build.sales_order
    co.sales_order_item_idx = build.sales_order_item_idx
    co.top_item = build.top_item
    co.top_bom = build.top_bom
    co.orientation = build.orientation
    co.snapshot = snapshot_name
    if hasattr(co, "flat_bom_status"):
        # Candidate/prod data has historically carried "Pending" here, but
        # the DocType now allows only Queued/Running/Complete/Failed. Set a
        # valid value before later package sync saves the CO.
        co.flat_bom_status = "Queued"
    for code in selected_codes:
        option_name = frappe.db.get_value("InductOne Configuration Option", {"option_code": code}, "name")
        option = frappe.get_doc("InductOne Configuration Option", option_name)
        co.append("selected_options", {
            "option": option.name,
            "option_code": option.option_code,
            "option_name": option.option_name,
            "option_category": getattr(option, "option_category", None),
            "builder_description": getattr(option, "builder_description", None),
            "internal_notes": getattr(option, "internal_notes", None),
        })
    co.insert(ignore_permissions=True, ignore_mandatory=True)
    build.latest_config_order = co.name
    build.config_order_rev = 1
    build.save(ignore_permissions=True)
    frappe.db.commit()
    return co


def create_minimal_generated_package(frappe, build, co_name: str, snapshot_name: str, generate_now):
    package = frappe.get_doc({
        "doctype": "BOM Export Package",
        "bom": build.top_bom,
        "source_mode": "Configured Build",
        "inductone_build": build.name,
        "configuration_order": co_name,
        "configured_snapshot": snapshot_name,
        "builder_supplier": build.builder_supplier,
        "include_pdf": 0,
        "include_stl": 0,
        # At least one output type is required by the package generator. Keep
        # DXF enabled but attachment collection disabled so this lifecycle
        # smoke validates the package state machine without duplicating the
        # heavyweight part-documentation ZIP validation suite.
        "include_dxf": 1,
        "include_step": 0,
        "include_qty": 1,
        "include_item_attachments": 0,
        "include_bom_attachments": 0,
        "explosion_mode": "Follow Explicit Child BOM Links",
        "status": "Draft",
        "run_log": "Candidate CSA lifecycle smoke minimal package.",
    })
    package.insert(ignore_permissions=True, ignore_mandatory=True)
    frappe.db.commit()
    generate_now(package.name)
    package = frappe.get_doc("BOM Export Package", package.name)
    build.latest_bom_export_package = package.name
    build.save(ignore_permissions=True)
    co = frappe.get_doc("InductOne Configuration Order", co_name)
    if hasattr(co, "bom_export_package"):
        co.bom_export_package = package.name
    co.save(ignore_permissions=True)
    frappe.db.commit()
    return package


def create_filled_completion_workbook(frappe, template_file_url: str, system_serial: str, build_name: str, timestamp: str) -> dict:
    from frappe.utils.file_manager import save_file

    if not template_file_url:
        raise RuntimeError("release did not return serial_template_file_url")
    workbook_bytes = read_file_url_bytes(frappe, template_file_url)
    wb = load_workbook(io.BytesIO(workbook_bytes))
    ws = wb["Builder Input"]
    for row_idx in range(1, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value
        if label is None:
            continue
        label_text = str(label).strip()
        if not label_text or label_text.upper().startswith("SECTION "):
            continue
        if label_text == "InductOne Serial Number (IND-####)":
            ws.cell(row=row_idx, column=2).value = system_serial
        elif label_text == "Build Date":
            ws.cell(row=row_idx, column=2).value = datetime.now(timezone.utc).date().isoformat()
        elif label_text == "Builder Organization":
            ws.cell(row=row_idx, column=2).value = "Motion Controls"
        elif label_text == "Builder Point of Contact":
            ws.cell(row=row_idx, column=2).value = "Candidate Lifecycle Smoke"
        elif label_text == "Builder Point of Contact Email":
            ws.cell(row=row_idx, column=2).value = "motion.builder@plusonerobotics.com"
        elif label_text == "Builder Signature (Typed Full Name)":
            ws.cell(row=row_idx, column=2).value = "Candidate Lifecycle Smoke"
        elif label_text == "Date":
            ws.cell(row=row_idx, column=2).value = datetime.now(timezone.utc).date().isoformat()
        elif label_text == "I confirm all entries are accurate (YES/NO)":
            ws.cell(row=row_idx, column=2).value = "YES"
        else:
            safe = "".join(ch for ch in label_text.upper() if ch.isalnum())[:20] or "COMP"
            ws.cell(row=row_idx, column=2).value = f"{system_serial}-{safe}"

    out = io.BytesIO()
    wb.save(out)
    file_name = f"{build_name}_completed_builder_workbook_{timestamp}.xlsx"
    saved = save_file(
        file_name,
        out.getvalue(),
        dt="InductOne Build",
        dn=build_name,
        is_private=1,
    )
    frappe.db.commit()
    return {"file_name": file_name, "file_url": saved.file_url, "file_doc": saved.name}


def read_file_url_bytes(frappe, file_url: str) -> bytes:
    from frappe.utils.file_manager import get_file_path

    file_name = frappe.db.get_value("File", {"file_url": file_url}, "name")
    if not file_name:
        raise FileNotFoundError(f"No File record for {file_url}")
    path = get_file_path(file_url)
    return Path(path).read_bytes()


def summarize_doc(doc, fields: list[str]) -> dict:
    return {field: doc.get(field) for field in fields if hasattr(doc, field)}


def add_step(payload: dict, label: str, passed: bool, details: dict | list | str | None = None) -> None:
    payload["steps"].append({
        "label": label,
        "passed": bool(passed),
        "summary": summarize_details(details),
        "details": details,
    })
    print("PASS" if passed else "FAIL", label, summarize_details(details), flush=True)


def assert_last_step(payload: dict) -> None:
    if not payload["steps"][-1]["passed"]:
        raise RuntimeError(f"Lifecycle smoke failed at step: {payload['steps'][-1]['label']}")


def summarize_details(details) -> str:
    if details is None:
        return ""
    if isinstance(details, str):
        return details[:300]
    text = json.dumps(details, default=str, sort_keys=True)
    return text[:500]


if __name__ == "__main__":
    sys.exit(main())
