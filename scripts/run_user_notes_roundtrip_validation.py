#!/usr/bin/env python3
"""Candidate-only round-trip validation for BOM Item per-line User Notes.

Validation snapshots are generated on scratch builds only; a real Build's
snapshot history is a clean audit trail.
"""

from __future__ import annotations

import argparse
import json
import tempfile
import time
from pathlib import Path

import frappe


DEFAULT_EVIDENCE_DIR = "/mnt/c/hub/frappe-sandbox/validation-evidence"


def record(results, name, ok, **extra):
    row = {"check": name, "ok": bool(ok), **extra}
    results.append(row)
    print(("PASS" if ok else "FAIL"), name, extra)
    return ok


def find_validation_context():
    builds = frappe.get_all(
        "InductOne Build",
        filters={"top_bom": ["is", "set"], "sales_order": ["is", "set"]},
        fields=["name", "sales_order", "top_bom", "top_item", "builder_supplier"],
        order_by="modified desc",
        limit=10,
    )
    if not builds:
        raise RuntimeError("No InductOne Build with top_bom and sales_order found.")

    from inductone_tools.bom_export import explode_bom_tree_structured

    for build in builds:
        rows = explode_bom_tree_structured(
            root_bom=build.top_bom,
            explosion_mode="Follow Explicit Child BOM Links",
            max_depth=0,
            include_qty=True,
        )
        source_rows = [r for r in rows if r.get("source_bom_item")]
        if source_rows:
            target = next((r for r in source_rows if r.get("is_leaf")), source_rows[0])
            leaf_codes = sorted({r["item_code"] for r in rows if r.get("is_leaf") and r.get("item_code")})
            return build, rows, target, leaf_codes
    raise RuntimeError("No explodable BOM Item rows found under available InductOne Builds.")


def create_scratch_build(source_build):
    name = f"{source_build.name}-USER-NOTES-SCRATCH-{int(time.time())}"
    doc = frappe.new_doc("InductOne Build")
    doc.name = name
    doc.sales_order = source_build.sales_order
    doc.top_item = source_build.top_item
    doc.top_bom = source_build.top_bom
    doc.builder_supplier = source_build.builder_supplier
    doc.customer_project_label = f"User notes round-trip scratch cloned from {source_build.name}"
    doc.builder_po_reference = "User notes validation scratch"
    doc.build_status = "DRAFT"
    doc.completion_status = "Open"
    doc.insert(ignore_permissions=True, ignore_mandatory=True)
    frappe.db.commit()
    return doc


def create_snapshot(build, leaf_codes):
    snap = frappe.get_doc({
        "doctype": "Configured BOM Snapshot",
        "sales_order": build.sales_order,
        "inductone_build": build.name,
        "top_item": build.top_item,
        "top_bom": build.top_bom,
        "snapshot_rev": int(time.time()),
        "generated_at": frappe.utils.now_datetime(),
    })
    for code in leaf_codes:
        snap.append("lines", {
            "item_code": code,
            "included": 1,
            "qty": 1,
            "level": 1,
            "rule_reason": "User Notes round-trip validation",
        })
    snap.insert(ignore_permissions=True)
    return snap.name


def create_package(build, snapshot_name):
    package = frappe.get_doc({
        "doctype": "BOM Export Package",
        "bom": build.top_bom,
        "source_mode": "Configured Build",
        "inductone_build": build.name,
        "configured_snapshot": snapshot_name,
        "builder_supplier": build.builder_supplier,
        "include_pdf": 1,
        "include_qty": 1,
        "include_item_attachments": 0,
        "include_bom_attachments": 0,
        "explosion_mode": "Follow Explicit Child BOM Links",
        "status": "Draft",
    })
    package.insert(ignore_permissions=True)
    return package.name


def workbook_contains(snapshot_name, expected):
    from inductone_tools.snapshot.hierarchy import _render_hierarchy_workbook
    from openpyxl import load_workbook

    snap = frappe.get_doc("Configured BOM Snapshot", snapshot_name)
    data = _render_hierarchy_workbook(snap)
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as fh:
        fh.write(data)
        path = Path(fh.name)
    try:
        wb = load_workbook(path)
        ws = wb.active
        headers = [ws.cell(row=9, column=c).value for c in range(1, ws.max_column + 1)]
        values = [
            ws.cell(row=r, column=c).value
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        return "User Notes" in headers and expected in values, headers
    finally:
        path.unlink(missing_ok=True)


def diff_detects_user_note_change(target_row, note_a, note_b):
    from inductone_tools.snapshot_diff.engine import USER_NOTES_CHANGED, diff_snapshots
    from inductone_tools.snapshot_diff.schema import SnapshotNode

    def node(suffix, note):
        return SnapshotNode(
            node_id=f"n-{suffix}",
            parent_node_id=None,
            bom_level=int(target_row.get("bom_level") or 1),
            item_code=target_row["item_code"],
            item_name=target_row.get("item_name") or target_row["item_code"],
            item_group="Validation",
            description=target_row.get("description") or "",
            qty=float(target_row.get("qty") or 1),
            uom=target_row.get("uom") or "",
            bom_used=target_row.get("bom_used") or "",
            node_type=target_row.get("node_type") or "Leaf",
            is_leaf=bool(target_row.get("is_leaf")),
            effect_origin="BASELINE",
            source_option_code="",
            excluded=False,
            source_bom=target_row.get("source_bom") or "",
            balloon_numbers=target_row.get("balloon_numbers") or "",
            electrical_unit=target_row.get("electrical_unit") or "",
            source_electrical_bom_rev=target_row.get("source_electrical_bom_rev") or "",
            user_notes=note,
        )

    result = diff_snapshots([node("a", note_a)], [node("b", note_b)], "A", "B", include_unchanged=True)
    line = result.lines[0] if result.lines else None
    return bool(line and USER_NOTES_CHANGED in line.categories), (line.categories if line else [])


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--site", required=True)
    parser.add_argument("--sites-path", required=True)
    parser.add_argument("--evidence-dir", default=DEFAULT_EVIDENCE_DIR)
    args = parser.parse_args()

    frappe.init(site=args.site, sites_path=args.sites_path)
    frappe.connect()
    frappe.set_user("Administrator")

    results = []
    created_snapshot = None
    created_package = None
    created_scratch_build = None
    original_values = {}
    note = f"USER NOTES ROUNDTRIP {int(time.time())}"
    note_changed = f"{note} CHANGED"

    try:
        source_build, baseline_rows, target, leaf_codes = find_validation_context()
        build = create_scratch_build(source_build)
        created_scratch_build = build.name
        record(results, "scratch_build_created_for_validation_snapshot", created_scratch_build != source_build.name, source_build=source_build.name, scratch_build=created_scratch_build)
        source_bom_item = target["source_bom_item"]
        original_doc = frappe.get_doc("BOM Item", source_bom_item)
        original_values = {
            "custom_user_notes": getattr(original_doc, "custom_user_notes", None),
            "custom_electrical_unit": getattr(original_doc, "custom_electrical_unit", None),
            "custom_source_electrical_bom_rev": getattr(original_doc, "custom_source_electrical_bom_rev", None),
        }

        frappe.db.set_value("BOM Item", source_bom_item, {
            "custom_user_notes": note,
            "custom_electrical_unit": "E-UNIT-VALIDATION",
            "custom_source_electrical_bom_rev": "REV-VALIDATION",
        })
        frappe.db.commit()

        from inductone_tools.bom_export import (
            explode_bom_tree_structured,
            update_results_and_missing_summary,
        )
        from inductone_tools.snapshot.hierarchy import populate_snapshot_hierarchy

        rows_after = explode_bom_tree_structured(build.top_bom, "Follow Explicit Child BOM Links", 0, True)
        matching_rows = [r for r in rows_after if r.get("source_bom_item") == source_bom_item]
        row = matching_rows[0] if matching_rows else {}
        record(results, "explode_structured_carries_user_notes", row.get("user_notes") == note, row=row)
        record(results, "explode_structured_regression_metadata", row.get("electrical_unit") == "E-UNIT-VALIDATION" and row.get("description") is not None and row.get("qty") is not None)

        created_snapshot = create_snapshot(build, leaf_codes)
        populate_result = populate_snapshot_hierarchy(created_snapshot)
        h_rows = frappe.get_all(
            "Configured BOM Snapshot Hierarchy",
            filters={"parent": created_snapshot, "source_bom_item": source_bom_item},
            fields=["name", "item_code", "electrical_unit", "source_electrical_bom_rev", "user_notes", "description", "qty"],
        )
        h_row = h_rows[0] if h_rows else {}
        record(results, "snapshot_hierarchy_carries_user_notes", h_row.get("user_notes") == note, hierarchy_row=h_row, populate_result=populate_result)

        created_package = create_package(build, created_snapshot)
        package = frappe.get_doc("BOM Export Package", created_package)
        missing = update_results_and_missing_summary(created_package, package, rows_after, {}, [".pdf"])
        p_rows = frappe.get_all(
            "BOM Export Package Item",
            filters={"parent": created_package, "item_code": target["item_code"]},
            fields=["name", "item_code", "user_notes", "qty"],
        )
        p_row = next((r for r in p_rows if r.get("user_notes") == note), p_rows[0] if p_rows else {})
        record(results, "bom_export_package_item_carries_user_notes", p_row.get("user_notes") == note, package_row=p_row, missing_count=len(missing))

        ok_wb, headers = workbook_contains(created_snapshot, note)
        record(results, "hierarchy_workbook_has_user_notes_column_and_value", ok_wb, headers=headers)

        ok_diff, categories = diff_detects_user_note_change(row, note, note_changed)
        record(results, "snapshot_diff_detects_user_notes_change", ok_diff, categories=categories)

        no_note_rows = [r for r in rows_after if not r.get("user_notes")]
        record(results, "blank_user_notes_default_empty_no_error", bool(no_note_rows), blank_row_count=len(no_note_rows))

    except Exception as exc:
        record(results, "unexpected_exception", False, exception=exc.__class__.__name__, message=str(exc))
    finally:
        if source_bom_item := (target.get("source_bom_item") if "target" in locals() else None):
            frappe.db.set_value("BOM Item", source_bom_item, original_values)
        if created_package:
            frappe.db.delete("BOM Export Package Item", {"parent": created_package})
            frappe.db.delete("BOM Export Package", {"name": created_package})
        if created_snapshot:
            frappe.db.delete("Configured BOM Snapshot Hierarchy", {"parent": created_snapshot})
            frappe.db.delete("Configured BOM Snapshot Item", {"parent": created_snapshot})
            frappe.db.delete("Configured BOM Snapshot", {"name": created_snapshot})
        if created_scratch_build:
            frappe.db.delete("InductOne Build", {"name": created_scratch_build})
        frappe.db.commit()
        frappe.destroy()

    ok = all(r["ok"] for r in results)
    evidence = {
        "site": args.site,
        "ok": ok,
        "results": results,
    }
    out_dir = Path(args.evidence_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    out = out_dir / f"user_notes_roundtrip_validation_{int(time.time())}.json"
    out.write_text(json.dumps(evidence, indent=2, default=str), encoding="utf-8")
    print(f"Evidence: {out}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
